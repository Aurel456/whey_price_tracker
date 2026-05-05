"""
HSN Whey Price Tracker
======================
Lance ce script pour mettre à jour les prix du jour dans whey_prices.xlsx
et régénérer le dashboard HTML whey_dashboard.html.

Usage: python hsn_tracker.py

Dépendances: pip install playwright openpyxl
             python -m playwright install chromium
"""

import asyncio
import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

# ── Configuration ─────────────────────────────────────────────────────────────
EXCEL_PATH       = Path(__file__).parent / "whey_prices.xlsx"
DESCRIPTIONS_PATH = Path(__file__).parent / "descriptions.json"
TAGS_PATH        = Path(__file__).parent / "tags.json"
ERROR_LOG_PATH   = Path(__file__).parent / "errors.log"

# Sanity checks : un whey doit avoir un % de protéine entre ces bornes
PROT_MIN_PCT = 50.0
PROT_MAX_PCT = 95.0
RETRY_ATTEMPTS = 1     # nb de retentatives en cas d'échec scraping
RETRY_DELAY_MS = 2_000

CATEGORY_URLS = [
    "https://www.hsnstore.fr/nutrition-sportive/proteines/whey",
    "https://www.hsnstore.fr/nutrition-sportive/proteines/whey?p=2",
    "https://www.hsnstore.fr/nutrition-sportive/proteines",
]

# URLs supplémentaires à toujours inclure (même si non listées en catégorie)
EXTRA_URLS = [
    "https://www.hsnstore.fr/marques/sport-series/evobasic-whey",
    "https://www.hsnstore.fr/marques/sport-series/evolate-2-0-sans-edulcorants-whey-isolate-cfm",
    "https://www.hsnstore.fr/marques/raw-series/isolat-de-proteine-de-lait",
    "https://www.hsnstore.fr/marques/sport-series/evowhey-protein-sans-edulcorants",
    "https://www.hsnstore.fr/marques/raw-series/isolat-de-proteine-hydrolysee-de-clear-whey",
    "https://www.hsnstore.fr/marques/raw-series/100-isolat-de-proteine-de-lactoserum-hydrolyse",
    "https://www.hsnstore.fr/marques/sport-series/evolate-2-0-whey-isolate-cfm",
    "https://www.hsnstore.fr/marques/raw-series/100-whey-protein-isolate",
]

# Oméga-3 (capsules, métriques différentes : €/g EPA+DHA)
OMEGA3_URLS = [
    "https://www.hsnstore.fr/marques/essential-series/premium-omega-3-tg-ifos-1000mg",
    "https://www.hsnstore.fr/marques/essential-series/omega-3-huile-de-poisson-1000mg",
    "https://www.hsnstore.fr/marques/essential-series/ultra-omega-3-tg-ifos-1000mg",
]

# Créatine (poudre, métrique : €/kg de créatine pure)
CREATINE_URLS = [
    "https://www.hsnstore.fr/marques/raw-series/creatine-excell-100-creapure-en-poudre",
    "https://www.hsnstore.fr/marques/raw-series/creatine-monohydratee-en-poudre-200-mesh",
    "https://www.hsnstore.fr/marques/raw-series/monohydrate-de-creatine-ultra-fin-en-poudre-500-mesh",
]

PAGE_TIMEOUT = 30_000
CLICK_WAIT   = 700
CONCURRENCY  = 4

PORT_RE = re.compile(
    r'(?:DDM:\s*([\d/]+)\s*\|)?\s*PORT\.:\s*(\d+)'
    r'(?:\s*\|\s*COÛT/PORT\.:\s*([\d,\s\xa0]+€))?'
    r'(?:\s*\|\s*PX/KG:\s*([\d,\s\xa0]+€))?'
)

# ── Logging d'erreurs ─────────────────────────────────────────────────────────
def log_error(url: str, reason: str) -> None:
    """Append une ligne dans errors.log avec timestamp + url + raison."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with ERROR_LOG_PATH.open("a", encoding="utf-8") as f:
        f.write(f"[{ts}] {url}\n   {reason}\n")


# ── Excel styles ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATE_FILL   = PatternFill("solid", fgColor="D9E1F2")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
CENT_ALIGN  = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
HEADERS    = [
    "Date", "Produit", "URL", "Taille", "Prix (€)",
    "Portions", "Coût/portion (€)", "Prix/kg (€)", "DDM",
    "Description courte", "Mots-clés",
    # Colonnes nutrition / coût protéine
    "Prix/kg protéine (€)", "Coût/30g protéine (€)",
    "Protéines (g/100g)", "Énergie (kcal/100g)",
    "Glucides (g/100g)", "Lipides (g/100g)", "Sel (g/100g)",
    "Leucine (mg/100g)", "Isoleucine (mg/100g)", "Valine (mg/100g)",
    "Ingrédients", "Profil AA (JSON)",
    # Score qualitatif : seuil anabolique = 3g leucine
    "Coût/3g leucine (€)",
    # Catégorie : Whey / Aliments enrichis / Autres (déduite du % de protéine)
    "Catégorie",
    # Type de produit : whey / omega3 / creatine (déduit de l'URL)
    "Type produit",
    # Champs spécifiques omega-3 et créatine (vides pour les whey)
    "EPA (mg/dose)", "DHA (mg/dose)", "Coût/g EPA+DHA (€)",
    "Créatine (g/dose)", "Coût/kg créatine (€)",
    # Disponibilité (False si rupture de stock détectée au scraping)
    "En stock",
]
COL_WIDTHS = [12, 45, 50, 12, 12, 10, 18, 15, 12, 60, 50,
              18, 18, 16, 18, 16, 16, 14, 16, 16, 14, 80, 60,
              18, 18,
              14, 14, 14, 18, 14, 18,
              10]


# ── Excel helpers ─────────────────────────────────────────────────────────────
def _style(cell, fill=None, font=None, alignment=None, border=THIN_BORDER):
    if fill:      cell.fill = fill
    if font:      cell.font = font
    if alignment: cell.alignment = alignment
    if border:    cell.border = border


def init_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Historique"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        c = ws.cell(row=1, column=col, value=h)
        _style(c, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENT_ALIGN)
        ws.column_dimensions[get_column_letter(col)].width = w
    wb.save(EXCEL_PATH)
    return wb


def load_or_create_workbook():
    if not EXCEL_PATH.exists():
        return init_workbook()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Historique"]
    existing = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if len(existing) < len(HEADERS):
        for col, h in enumerate(HEADERS[len(existing):], start=len(existing) + 1):
            c = ws.cell(row=1, column=col, value=h)
            _style(c, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENT_ALIGN)
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS[col - 1]
        wb.save(EXCEL_PATH)
    return wb


def _clean(val):
    if val is None:
        return None
    return str(val).replace("\xa0", " ").replace(",", ".").replace("€", "").strip()


def _to_float(val):
    c = _clean(val)
    if not c:
        return None
    try:
        return float(c)
    except ValueError:
        return None


def append_rows(rows: list):
    wb = load_or_create_workbook()
    ws = wb["Historique"]
    today = date.today().isoformat()
    next_row = ws.max_row + 1
    for i, r in enumerate(rows):
        aa = r.get("amino_acids") or {}
        data = [
            today, r.get("name", ""), r.get("url", ""), r.get("size", ""),
            _to_float(r.get("price")),
            _to_float(r.get("portions")) if r.get("portions") else None,
            _to_float(r.get("cout_portion")),
            _to_float(r.get("px_kg")),
            r.get("ddm", ""),
            r.get("desc_short", ""),
            r.get("keywords", ""),
            # Coût protéine + nutrition
            r.get("px_kg_proteine"),
            r.get("cout_30g_proteine"),
            r.get("proteines_100g"),
            r.get("energie_kcal_100g"),
            r.get("glucides_100g"),
            r.get("lipides_100g"),
            r.get("sel_100g"),
            aa.get("L-Leucine") or aa.get("Leucine"),
            aa.get("L-Isoleucine") or aa.get("Isoleucine"),
            aa.get("L-Valine") or aa.get("Valine"),
            r.get("ingredients", ""),
            json.dumps(aa, ensure_ascii=False) if aa else "",
            r.get("cout_3g_leucine"),
            r.get("categorie", ""),
            r.get("type_produit", "whey"),
            r.get("epa_mg_dose"),
            r.get("dha_mg_dose"),
            r.get("cout_g_epa_dha"),
            r.get("creatine_g_dose"),
            r.get("cout_kg_creatine"),
            r.get("en_stock", True),
        ]
        row_fill = ALT_FILL if i % 2 == 0 else None
        for col, val in enumerate(data, start=1):
            c = ws.cell(row=next_row, column=col, value=val)
            align = LEFT_ALIGN if col in (2, 3) else CENT_ALIGN
            _style(c, fill=row_fill, alignment=align)
            if col == 1:
                c.fill = DATE_FILL
                c.font = Font(bold=True, name="Arial", size=9)
            else:
                c.font = Font(name="Arial", size=9)
        next_row += 1
    wb.save(EXCEL_PATH)


# ── Scraping helpers ──────────────────────────────────────────────────────────
def parse_port_line(text: str) -> dict:
    for line in text.split("\n"):
        m = PORT_RE.search(line)
        if m:
            return {
                "ddm":          m.group(1),
                "portions":     m.group(2),
                "cout_portion": m.group(3),
                "px_kg":        m.group(4),
            }
    return {}


def extract_spconfig(page_source: str):
    m = re.search(
        r'initConfigurableOptions\s*\(\s*[\'"][\d]+[\'"]\s*,\s*(\{)', page_source
    )
    if not m:
        return None
    start = m.start(1)
    depth, end = 0, start
    for i in range(start, min(len(page_source), start + 300_000)):
        if page_source[i] == '{':
            depth += 1
        elif page_source[i] == '}':
            depth -= 1
            if depth == 0:
                end = i
                break
    try:
        return json.loads(page_source[start:end + 1])
    except json.JSONDecodeError:
        return None


def build_option_price_map(spconfig: dict) -> dict:
    op = spconfig.get("optionPrices", {})
    mapping = {}
    for attr in spconfig.get("attributes", {}).values():
        for opt in attr.get("options", []):
            opt_id = str(opt.get("id", ""))
            for pid in opt.get("products", []):
                pid_s = str(pid)
                if pid_s in op:
                    fp = op[pid_s].get("finalPrice", {}).get("amount")
                    if fp is not None:
                        mapping[opt_id] = fp
    return mapping


# ── Extraction nutrition / AA / ingrédients ───────────────────────────────────
_NUM_RE = re.compile(r'([\d,\.]+)')
_KCAL_RE = re.compile(r'([\d,\.]+)\s*Kcal', re.IGNORECASE)


def _parse_num(text):
    if not text:
        return None
    t = str(text).strip()
    m = _KCAL_RE.search(t)  # priorité aux Kcal pour la valeur énergétique
    if m:
        return float(m.group(1).replace(",", "."))
    m = _NUM_RE.search(t)
    if m:
        return float(m.group(1).replace(",", "."))
    return None


_TABLES_JS = r"""() => {
    const findHeading = (t) => {
        let p = t.parentElement;
        while (p) {
            const hs = p.querySelectorAll(':scope h1, :scope h2, :scope h3, :scope h4');
            let last = null;
            for (const h of hs) {
                if (t.compareDocumentPosition(h) & Node.DOCUMENT_POSITION_PRECEDING) last = h;
            }
            if (last) return last.innerText.trim();
            p = p.parentElement;
        }
        return '';
    };
    return Array.from(document.querySelectorAll('table')).map(t => ({
        heading: findHeading(t),
        rows: Array.from(t.rows).map(r =>
            Array.from(r.cells).map(c => c.innerText.trim())
        ),
    }));
}"""

_INGREDIENTS_JS = r"""() => {
    const headers = Array.from(document.querySelectorAll('p, h2, h3, h4, span, strong'))
        .filter(el => el.textContent.trim() === 'Ingrédients');
    for (const h of headers) {
        let next = h.nextElementSibling;
        while (next && !next.innerText?.trim()) next = next.nextElementSibling;
        if (next && next.innerText.trim().length > 20) {
            return next.innerText.trim();
        }
    }
    return '';
}"""


def _parse_nutrition(tables):
    """Extrait les valeurs nutritionnelles depuis les tables HTML.

    - Pour whey : col[2] = pour 100g (priorité).
    - Pour omega3 / créatine (table à 2 colonnes) : col[1] = par dose/capsule.
    - Détecte aussi EPA / DHA (mg/dose) et créatine (g/dose).
    """
    out = {}
    for tbl in tables:
        h = tbl["heading"].lower()
        if "nutritionnel" not in h and "valeurs nutrit" not in h and "composition" not in h:
            continue
        for row in tbl["rows"][1:]:
            if len(row) < 2:
                continue
            label = row[0].lower()
            # Pour whey on prend col[2] (per 100g) ; sinon on retombe sur col[1]
            val = _parse_num(row[2]) if len(row) >= 3 else None
            if val is None:
                val = _parse_num(row[1])
            if val is None:
                continue
            # Macros (whey)
            if "valeur" in label or "énerg" in label or "energ" in label:
                out["energie_kcal_100g"] = val
            elif "protéin" in label or "protein" in label:
                out["proteines_100g"] = val
            elif "sucre" in label:
                out["sucres_100g"] = val
            elif "glucides" in label or "hydrates" in label or "carbon" in label:
                out["glucides_100g"] = val
            elif "satur" in label:
                out["lipides_satures_100g"] = val
            elif "graisse" in label or "lipide" in label:
                out["lipides_100g"] = val
            elif "sel" in label or "sodium" in label:
                out["sel_100g"] = val
            # Omega-3
            elif "epa" in label or "eicosapent" in label or "icosapent" in label:
                out["epa_mg_dose"] = val
            elif "dha" in label or "docosahexa" in label:
                out["dha_mg_dose"] = val
            # Créatine (g/dose) — convertit mg → g si l'unité est en milligrammes
            elif ("créatine" in label or "creatine" in label) and "kreatin" not in label:
                raw = row[2] if len(row) >= 3 else row[1]
                raw_l = str(raw).lower()
                if "mg" in raw_l and " g" not in raw_l.replace("mg", ""):
                    out["creatine_g_dose"] = round(val / 1000.0, 2)
                else:
                    out["creatine_g_dose"] = val
        break
    return out


def _parse_amino_acids(tables):
    """Extrait le profil AA (mg/100g) depuis les tables HTML."""
    out = {}
    for tbl in tables:
        h = tbl["heading"].lower()
        if "acide" not in h or "amin" not in h:
            continue
        for row in tbl["rows"][1:]:
            if len(row) == 1:
                m = re.match(r'(.+?)([\d\.,]+)\s*mg', row[0])
                if m:
                    out[m.group(1).strip()] = float(m.group(2).replace(",", "."))
            elif len(row) >= 2:
                m = re.search(r'([\d\.,]+)\s*mg', row[1])
                if m:
                    out[row[0].strip()] = float(m.group(1).replace(",", "."))
        break
    return out


async def extract_nutrition_data(page) -> dict:
    """Récupère nutrition /100g + acides aminés + ingrédients (1× par produit)."""
    # Scroll pour déclencher le lazy-load des tables
    for _ in range(6):
        await page.evaluate("window.scrollBy(0, 1200)")
        await page.wait_for_timeout(150)

    tables = await page.evaluate(_TABLES_JS)
    ingredients = await page.evaluate(_INGREDIENTS_JS) or ""
    return {
        "nutrition": _parse_nutrition(tables),
        "amino_acids": _parse_amino_acids(tables),
        "ingredients": ingredients,
    }


SWEETENER_LABELS = {
    "sucralose":      "Sucralose",
    "stevia":         "Stévia",
    "acesulfame_k":   "Acésulfame-K",
    "aspartame":      "Aspartame",
    "sans_edulcorant": "Sans édulcorant",
}

WHEY_TYPE_LABELS = {
    "isolat_cfm":  "Isolat CFM (filtration à froid)",
    "hydrolysat":  "Hydrolysat",
    "isolat":      "Isolat",
    "concentre":   "Concentré",
    "native":      "Native (lait cru)",
    "caseine":     "Caséine",
    "vegetal":     "Végétal",
    "mix":         "Mix protéines",
    "enzyme":      "Enzymes digestives",
}


def _detect_sweeteners(ingredients: str, name: str) -> list:
    """Détecte les édulcorants à partir des ingrédients et du nom du produit.

    Liste de tags possibles : sucralose, stevia, acesulfame_k, aspartame, sans_edulcorant.
    Plusieurs tags peuvent être renvoyés (ex : sucralose + stevia).
    """
    s = (ingredients or "").lower()
    name_l = (name or "").lower()
    out = []
    if "sucralose" in s:
        out.append("sucralose")
    if "stévi" in s or "stevi" in s or "glycoside" in s:
        out.append("stevia")
    if "acésulf" in s or "acesulf" in s:
        out.append("acesulfame_k")
    if "aspartame" in s:
        out.append("aspartame")
    if not out:
        # Pas d'édulcorant détecté → considéré "sans édulcorant" (sauf si le mot
        # "édulcorant" apparaît mais qu'on n'a pas matché — alors on s'abstient).
        if "édulcorant" not in s and "edulcorant" not in s:
            out.append("sans_edulcorant")
        elif "sans édulcorant" in name_l or "sans edulcorant" in name_l:
            out.append("sans_edulcorant")
    return out


def _detect_whey_type(ingredients: str, name: str) -> list:
    """Détecte le type de protéine (isolat / hydrolysat / CFM / native / végétal …).

    Combine ingrédients + nom du produit. Plusieurs tags possibles (ex : isolat_cfm + hydrolysat).
    """
    s = (ingredients or "").lower()
    name_l = (name or "").lower()
    combined = f"{s} {name_l}"
    out = []

    is_cfm = (
        "cfm" in combined
        or "cross flow" in combined
        or "cross-flow" in combined
        or "extrait à froid" in combined
        or "extraction à froid" in combined
        or "filtration à froid" in combined
        or "microfiltration" in combined
    )
    if is_cfm:
        out.append("isolat_cfm")
    if "hydrolys" in combined:
        out.append("hydrolysat")
    if "native" in combined or "natif" in combined:
        out.append("native")
    if "caséine" in combined or "caseine" in combined or "peptopro" in combined:
        out.append("caseine")
    # Végétal : soja / pois / riz / chanvre / vegan
    if (
        "soja" in s or "vegan" in name_l or "vegan" in s
        or "pois" in s or "chanvre" in s
        or "riz" in s and "crème de riz" in name_l
    ):
        out.append("vegetal")
    if "concentré" in combined or "concentrate" in combined or "concentree" in combined:
        out.append("concentre")
    if ("isolat" in combined or "isolate" in combined) and "isolat_cfm" not in out:
        out.append("isolat")
    # Mix si plusieurs sources de protéines distinctes (très spécifique)
    if "mix" in name_l and "protéines" in name_l:
        out.append("mix")
    # Enzymes digestives : marque DigeZyme® ou mention explicite
    if "digezyme" in combined or "enzyme" in combined:
        out.append("enzyme")
    return out


OMEGA3_TYPE_LABELS = {
    "form_tg":   "Triglycéride (TG)",
    "form_ee":   "Ester éthylique (EE)",
    "ifos":      "IFOS certifié",
}

CREATINE_TYPE_LABELS = {
    "creapure":   "Creapure®",
    "monohydrate": "Monohydrate",
    "mesh_100":   "100 mesh",
    "mesh_200":   "200 mesh",
    "mesh_500":   "500 mesh (ultra-fine)",
}


def _detect_product_type(url: str) -> str:
    """Déduit le type de produit (whey / omega3 / creatine) depuis l'URL."""
    if not url:
        return "whey"
    u = url.lower()
    if "omega-3" in u or "/omega" in u or "huile-de-poisson" in u:
        return "omega3"
    if "creatine" in u or "créatine" in u or "creapure" in u or "monohydrate" in u:
        return "creatine"
    return "whey"


def _detect_omega3_tags(name: str, ingredients: str) -> list:
    s = f"{(name or '').lower()} {(ingredients or '').lower()}"
    out = []
    if " tg" in s or "(tg)" in s or "triglycéride" in s or "triglyceride" in s:
        out.append("form_tg")
    if " ee" in s or "(ee)" in s or "ester éthyl" in s or "ester ethyl" in s:
        out.append("form_ee")
    if "ifos" in s:
        out.append("ifos")
    return out


def _detect_creatine_tags(name: str, ingredients: str) -> list:
    s = f"{(name or '').lower()} {(ingredients or '').lower()}"
    out = []
    if "creapure" in s:
        out.append("creapure")
    if "monohydrat" in s:
        out.append("monohydrate")
    if "100 mesh" in s or "100mesh" in s:
        out.append("mesh_100")
    if "200 mesh" in s or "200mesh" in s:
        out.append("mesh_200")
    if "500 mesh" in s or "500mesh" in s or "ultra-fin" in s or "ultra fin" in s:
        out.append("mesh_500")
    return out


CAPS_RE = re.compile(r'(\d+)\s*(?:caps?|capsule|gélule|gelules?|softgel|comprim|tabl)', re.IGNORECASE)


def _parse_size_caps(label: str):
    """Parse '60 capsules' / '120 caps' / '90 softgels' → nombre de capsules."""
    if not label:
        return None
    m = CAPS_RE.search(label)
    return int(m.group(1)) if m else None


def load_manual_tags() -> dict:
    """Charge tags.json (annotations utilisateur) si présent.

    Format attendu :
        {
          "PRODUIT NAME": {
            "labels": ["favori", "à tester"],
            "note": "texte libre"
          }
        }
    """
    if not TAGS_PATH.exists():
        return {}
    try:
        return json.loads(TAGS_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        print(f"  [warn] tags.json illisible : {e}")
        return {}


def _detect_category(prot_per_100g) -> str:
    """Catégorise un produit selon sa teneur en protéines pour 100g.

    - Whey : >= 70% (whey isolat / hydrolysat / concentré pur)
    - Aliments enrichis : 30-70% (mix avec glucides/fibres : oats, smoothies, riz)
    - Autres : < 30% ou inconnu (à investiguer)
    """
    if prot_per_100g is None:
        return ""
    if prot_per_100g >= 70:
        return "Whey"
    if prot_per_100g >= 30:
        return "Aliments enrichis"
    return "Autres"


def _compute_protein_costs(px_kg, prot_per_100g, leucine_mg_per_100g=None):
    """Calcule prix/kg protéine, coût/30g protéine, et coût/3g leucine (seuil anabolique)."""
    if px_kg is None or not prot_per_100g:
        return None, None, None
    ratio = prot_per_100g / 100.0
    if ratio <= 0:
        return None, None, None
    px_kg_prot = round(px_kg / ratio, 2)
    cout_30g = round(px_kg_prot * 0.030, 3)
    # Coût pour atteindre 3g de leucine (seuil de stimulation MPS)
    cout_3g_leu = None
    if leucine_mg_per_100g and leucine_mg_per_100g > 0:
        # px_kg = €/kg produit ; leucine_mg/100g → leucine_g/kg = leucine_mg / 100
        leucine_g_per_kg = leucine_mg_per_100g / 100.0
        cout_3g_leu = round(px_kg / leucine_g_per_kg * 3.0, 3)
    return px_kg_prot, cout_30g, cout_3g_leu


async def dismiss_cookie_popup(page) -> None:
    try:
        await page.click('#didomi-notice-agree-button', timeout=3_000)
        await page.wait_for_timeout(600)
    except Exception:
        await page.evaluate("""() => {
            document.getElementById('didomi-host')?.remove();
            document.querySelector('.didomi-popup-backdrop')?.remove();
        }""")


async def get_product_urls(page, category_urls: list) -> list:
    seen, urls = set(), []
    for cat_url in category_urls:
        try:
            await page.goto(cat_url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
            try:
                await page.wait_for_selector('a.product-item-link, .product-item-info a', timeout=8_000)
            except PlaywrightTimeout:
                pass
            await dismiss_cookie_popup(page)
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(600)
            links = await page.evaluate("""() =>
                Array.from(document.querySelectorAll('a.product-item-link, .product-item-info a'))
                    .map(a => a.href)
                    .filter(h => h.includes('/marques/') ||
                                 (h.includes('/proteines/') && h.split('/').length > 5))
            """)
            added = 0
            for link in links:
                clean = link.split("?")[0].rstrip("/")
                if clean not in seen:
                    seen.add(clean)
                    urls.append(link.split("?")[0])
                    added += 1
            print(f"  {cat_url.split('fr/')[-1]:50} +{added}")
        except Exception as e:
            print(f"  Erreur {cat_url}: {e}")
    return urls


SIZE_EXCLUDE_RE = re.compile(r'monodose|pack', re.IGNORECASE)
SIZE_KG_RE = re.compile(r'(\d+(?:[\.,]\d+)?)\s*([Kk]?[Gg])')


def _parse_size_kg(label: str):
    """Parse '500g' / '1Kg' / '2Kg' → poids en kg (float)."""
    if not label:
        return None
    m = SIZE_KG_RE.match(label.strip())
    if not m:
        return None
    val = float(m.group(1).replace(",", "."))
    unit = m.group(2).lower()
    return val if unit.startswith("k") else val / 1000


def _enrich_row(row: dict, nutri: dict) -> dict:
    """Ajoute nutrition + acides aminés + métriques selon le type de produit."""
    n = nutri.get("nutrition", {})
    aa = nutri.get("amino_acids", {})
    row["amino_acids"] = aa
    row["ingredients"] = nutri.get("ingredients", "")
    row.update(n)

    ptype = _detect_product_type(row.get("url", ""))
    row["type_produit"] = ptype

    px_kg = _to_float(row.get("px_kg"))
    price = _to_float(row.get("price"))

    if ptype == "whey":
        leucine = aa.get("L-Leucine") or aa.get("Leucine")
        px_kg_prot, cout_30g, cout_3g_leu = _compute_protein_costs(
            px_kg, n.get("proteines_100g"), leucine
        )
        row["px_kg_proteine"]    = px_kg_prot
        row["cout_30g_proteine"] = cout_30g
        row["cout_3g_leucine"]   = cout_3g_leu
        row["categorie"]         = _detect_category(n.get("proteines_100g"))
        # Sanity check protéine
        prot = n.get("proteines_100g")
        if prot is not None and not (PROT_MIN_PCT <= prot <= PROT_MAX_PCT):
            log_error(
                row.get("url", "?"),
                f"Protéine suspecte : {prot}g/100g hors [{PROT_MIN_PCT};{PROT_MAX_PCT}] "
                f"(taille={row.get('size','?')}) — possible erreur de parsing"
            )

    elif ptype == "omega3":
        # €/g d'EPA+DHA = prix / (capsules_pack × (EPA+DHA mg/cap) / 1000)
        caps = _parse_size_caps(row.get("size", ""))
        epa = n.get("epa_mg_dose")
        dha = n.get("dha_mg_dose") or 0
        if epa is not None and caps and price:
            total_g = caps * (epa + dha) / 1000.0
            if total_g > 0:
                row["cout_g_epa_dha"] = round(price / total_g, 3)
        # px_kg synthétique : capsule 1000mg → 1g d'huile/cap
        # → 1 kg d'huile = 1000 caps. px_kg = price × 1000 / caps
        cap_mg = 1000  # par défaut HSN affiche "1000mg" dans le nom
        m = re.search(r'(\d+)\s*mg', (row.get("name") or ""), re.IGNORECASE)
        if m:
            cap_mg = float(m.group(1))
        if caps and price and cap_mg:
            px_kg_oil = price / (caps * cap_mg / 1_000_000.0)
            row["px_kg"] = f"{px_kg_oil:.2f} €"
        row["categorie"] = "Oméga-3"

    elif ptype == "creatine":
        # €/kg de créatine (≈ €/kg produit, monohydrate ~99% pur)
        size_kg = _parse_size_kg(row.get("size", ""))
        if size_kg and price:
            row["cout_kg_creatine"] = round(price / size_kg, 2)
        row["categorie"] = "Créatine"

    return row


_STOCK_CHECK_JS = r"""() => {
    // Check restreint à la zone produit principale (pas tout le body) pour
    // éviter qu'une variante OOS (ex: Pack) contamine la détection des
    // autres tailles. Sur Magento/HSN, l'état stock de la variante
    // SÉLECTIONNÉE se reflète dans .product-info-main / l'add-to-cart.
    const main = document.querySelector('.product-info-main')
              || document.querySelector('.product-info-price')
              || document.querySelector('[data-product-id]')
              || document.body;
    if (main.querySelector('.stock.unavailable')) return false;
    // Bouton "Prévenez-moi quand dispo" visible → OOS
    const alertBtn = main.querySelector(
        'button[id*="alert"], a[href*="alert"], .alert-stock, [data-action*="alert"]'
    );
    if (alertBtn && alertBtn.offsetParent !== null) return false;
    // Bouton ajouter au panier : visible et non-disabled → en stock
    const cart = main.querySelector(
        '#product-addtocart-button, button.tocart, button.action.tocart'
    );
    if (cart) {
        if (cart.disabled) return false;
        const style = window.getComputedStyle(cart);
        if (style.display === 'none' || style.visibility === 'hidden') return false;
        return true;
    }
    return true;
}"""


async def scrape_product(page, url: str) -> list:
    results = []
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
        try:
            await page.wait_for_selector('h1', timeout=10_000)
        except PlaywrightTimeout:
            pass
        await dismiss_cookie_popup(page)

        name = await page.evaluate(
            "document.querySelector('h1.page-title span, h1[itemprop=\"name\"]')"
            "?.innerText?.trim() || document.querySelector('h1')?.innerText?.trim() || ''"
        )

        html = await page.content()
        spconfig = extract_spconfig(html)
        option_price_map = build_option_price_map(spconfig) if spconfig else {}

        # Nutrition / AA / ingrédients : 1× par produit (identique sur toutes les tailles)
        nutri = await extract_nutrition_data(page)

        # Méthode legacy : inputs super_attribute (radio/checkbox)
        sizes = await page.evaluate("""() =>
            Array.from(document.querySelectorAll('input[name*="super_attribute"]')).map(i => ({
                kind: 'input',
                value: i.value,
                id: i.id,
                label: document.querySelector('label[for="' + i.id + '"]')?.innerText?.trim()
            })).filter(s => s.label)
        """)

        # Méthode actuelle (HSN ~2025+) : <select id="selectProductSimple"> avec
        # options du type "EVOCLEAR HYDRO 1Kg ANANAS" — on groupe par taille
        if not sizes:
            variants = await page.evaluate(r"""() =>
                Array.from(document.querySelectorAll('#selectProductSimple option, select.select-product option'))
                    .map(o => ({sku: o.value, text: (o.textContent || '').trim()}))
                    .filter(o => o.sku && o.text && !/Sélectionnez/i.test(o.text))
            """)
            size_re = re.compile(r'(\d+(?:[\.,]\d+)?\s*[Kk]?[Gg])')
            by_size = {}
            for v in variants:
                # Exclure les variantes Monodose / Pack au niveau du texte complet
                if SIZE_EXCLUDE_RE.search(v["text"]):
                    continue
                m = size_re.search(v["text"])
                if not m:
                    continue
                sz_label = m.group(1).strip().replace(" ", "")  # "1 Kg" → "1Kg"
                if sz_label not in by_size:
                    by_size[sz_label] = v["sku"]
            sizes = [
                {"kind": "select", "sku": sku, "label": sz}
                for sz, sku in by_size.items()
            ]

        # Exclure les tailles Monodose / Pack
        sizes = [s for s in sizes if not SIZE_EXCLUDE_RE.search(s["label"])]

        if not sizes:
            text = await page.evaluate("document.body.innerText")
            port_data = parse_port_line(text)
            price_raw = await page.evaluate(
                "document.querySelector('[data-price-type=\"finalPrice\"] .price')?.innerText || ''"
            )
            en_stock = await page.evaluate(_STOCK_CHECK_JS)
            row = {
                "name": name, "url": url, "size": "Unique",
                "price": _clean(price_raw), "en_stock": en_stock, **port_data,
            }
            results.append(_enrich_row(row, nutri))
            return results

        for sz in sizes:
            price_str = None
            if sz.get("kind") == "input":
                # Legacy : click sur label
                price_amount = option_price_map.get(sz["value"])
                try:
                    await page.click(f'label[for="{sz["id"]}"]')
                    await page.wait_for_timeout(CLICK_WAIT)
                except Exception:
                    pass
                if price_amount is not None:
                    price_str = f"{price_amount:.2f}"
            else:
                # Select : prix depuis spconfig.optionPrices (keyed by SKU)
                if spconfig:
                    op = spconfig.get("optionPrices", {})
                    fp = op.get(str(sz["sku"]), {}).get("finalPrice", {}).get("amount")
                    if fp is not None:
                        price_str = f"{fp:.2f}"
                # Sélectionne l'option pour que le DOM reflète l'état stock
                # de cette variante précise (sinon on lit l'état de la variante
                # par défaut pour toutes les tailles).
                for sel in ("#selectProductSimple", "select.select-product"):
                    try:
                        await page.select_option(sel, value=str(sz["sku"]))
                        await page.wait_for_timeout(CLICK_WAIT)
                        break
                    except Exception:
                        continue

            # Stock par variante (après sélection/click).
            en_stock = await page.evaluate(_STOCK_CHECK_JS)

            # Port data : seulement fiable après un click (legacy).
            # Pour les selects, on calcule px_kg nous-mêmes depuis prix/poids.
            if sz.get("kind") == "select":
                size_kg = _parse_size_kg(sz["label"])
                px_val = _to_float(price_str)
                port_data = {}
                if px_val is not None and size_kg:
                    px_kg_val = px_val / size_kg
                    port_data["px_kg"] = f"{px_kg_val:.2f} €"
            else:
                text = await page.evaluate("document.body.innerText")
                port_data = parse_port_line(text)

            row = {
                "name": name, "url": url, "size": sz["label"],
                "price": price_str, "en_stock": en_stock, **port_data,
            }
            results.append(_enrich_row(row, nutri))
            short = url.rsplit('/', 1)[-1][:28]
            stock_flag = "" if en_stock else " [OOS]"
            print(
                f"    [{short:28s}] {sz['label']:14} | {(price_str or '?'):>8} EUR | "
                f"PORT:{port_data.get('portions','?'):>4} | "
                f"PROT:{nutri['nutrition'].get('proteines_100g','?')}g | "
                f"PX/KG-PROT:{row.get('px_kg_proteine','?')}{stock_flag}"
            )

    except PlaywrightTimeout:
        print(f"  Timeout : {url}")
    except Exception as e:
        print(f"  Erreur : {url} — {e}")

    return results


# ── Dashboard HTML ────────────────────────────────────────────────────────────
def generate_dashboard(rows=None):
    """Regenerate whey_dashboard.html from current Excel data (or provided rows)."""
    if rows is None:
        wb = load_or_create_workbook()
        ws = wb["Historique"]
        hdrs = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = [
            dict(zip(hdrs, r))
            for r in ws.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in r)
        ]

    # Construit l'historique complet par (produit, taille) pour les tendances
    from collections import defaultdict
    history_by_key = defaultdict(list)
    latest = {}
    for r in rows:
        size = str(r.get("Taille", ""))
        if not size or "Pack" in size or "Monodose" in size:
            continue
        # Garde la ligne si on a au moins un prix (price OU px_kg). Pour les oméga
        # en capsules, px_kg est None mais le prix existe — il faut les conserver.
        if r.get("Prix (€)") is None and r.get("Prix/kg (€)") is None:
            continue
        key = (str(r.get("Produit", "")), size)
        rdate = str(r.get("Date", ""))
        history_by_key[key].append({
            "date": rdate,
            "pxkg": r.get("Prix/kg (€)"),
            "pxkgProt": r.get("Prix/kg protéine (€)"),
            "cout3leu": r.get("Cout/3g leucine (€)") or r.get("Coût/3g leucine (€)"),
            "coutGOmega": r.get("Coût/g EPA+DHA (€)") or r.get("Cout/g EPA+DHA (€)"),
            "coutKgCrea": r.get("Coût/kg créatine (€)") or r.get("Cout/kg créatine (€)"),
        })
        if key not in latest or rdate >= str(latest[key].get("Date", "")):
            latest[key] = r

    # Calcule moyenne historique (hors dernière date) + flag deal
    deal_meta = {}
    for key, hist in history_by_key.items():
        hist.sort(key=lambda x: x["date"])
        latest_date = hist[-1]["date"] if hist else None
        previous_pxp = [h["pxkgProt"] for h in hist if h["date"] != latest_date and h["pxkgProt"] is not None]
        avg = (sum(previous_pxp) / len(previous_pxp)) if previous_pxp else None
        cur = hist[-1]["pxkgProt"] if hist and hist[-1]["pxkgProt"] is not None else None
        is_deal = bool(avg and cur and cur < avg * 0.95)  # -5% sous la moyenne
        deal_meta[key] = {"avg": avg, "isDeal": is_deal, "histLen": len(hist)}

    manual_tags = load_manual_tags()

    chart_rows = []
    for r in latest.values():
        produit = str(r.get("Produit", ""))
        taille  = str(r.get("Taille", ""))
        url = str(r.get("URL", ""))
        key = (produit, taille)
        meta = deal_meta.get(key, {})
        ingr = r.get("Ingrédients") or ""
        # type_produit : si la colonne Excel n'existe pas (ancien historique), retomber sur l'URL
        ptype = str(r.get("Type produit") or _detect_product_type(url))
        if ptype == "whey":
            sweeteners = _detect_sweeteners(ingr, produit)
            whey_types = _detect_whey_type(ingr, produit)
            type_tags = []
        elif ptype == "omega3":
            sweeteners, whey_types = [], []
            type_tags = _detect_omega3_tags(produit, ingr)
        elif ptype == "creatine":
            sweeteners, whey_types = [], []
            type_tags = _detect_creatine_tags(produit, ingr)
        else:
            sweeteners, whey_types, type_tags = [], [], []
        manual = manual_tags.get(produit, {}) or {}
        chart_rows.append({
            "date":     str(r.get("Date", "")),
            "produit":  produit,
            "taille":   taille,
            "type":     ptype,
            "prix":     r.get("Prix (€)"),
            "pxkg":     r.get("Prix/kg (€)"),
            "cout":     r.get("Cout/portion (€)") or r.get("Coût/portion (€)"),
            "portions": r.get("Portions"),
            "ddm":      str(r.get("DDM") or ""),
            "url":      url,
            "pxkgProt": r.get("Prix/kg protéine (€)"),
            "cout30":   r.get("Cout/30g protéine (€)") or r.get("Coût/30g protéine (€)"),
            "prot":     r.get("Protéines (g/100g)"),
            "cout3leu": r.get("Cout/3g leucine (€)") or r.get("Coût/3g leucine (€)"),
            "epa":         r.get("EPA (mg/dose)"),
            "dha":         r.get("DHA (mg/dose)"),
            "coutGOmega":  r.get("Coût/g EPA+DHA (€)") or r.get("Cout/g EPA+DHA (€)"),
            "creaDose":    r.get("Créatine (g/dose)"),
            "coutKgCrea":  r.get("Coût/kg créatine (€)") or r.get("Cout/kg créatine (€)"),
            "avgPxkgProt": meta.get("avg"),
            "isDeal":      meta.get("isDeal", False),
            "histLen":     meta.get("histLen", 0),
            "categorie":   str(r.get("Catégorie") or _detect_category(r.get("Protéines (g/100g)"))),
            "sweeteners":  sweeteners,
            "wheyTypes":   whey_types,
            "typeTags":    type_tags,
            "ingredients": ingr,
            "labels":      list(manual.get("labels", []) or []),
            "note":        str(manual.get("note", "") or ""),
            "hasIngredients": bool(ingr.strip()) if isinstance(ingr, str) else False,
            "en_stock":    r.get("En stock"),
        })

    # Sérialise l'historique pour les courbes de tendance
    # Catégorie reprise depuis le snapshot le plus récent (chart_rows / latest)
    cat_by_key = {(c["produit"], c["taille"]): c["categorie"] for c in chart_rows}
    history_json_data = [
        {
            "produit":   k[0],
            "taille":    k[1],
            "categorie": cat_by_key.get(k, ""),
            "points":    v,
        }
        for k, v in history_by_key.items()
        if len(v) >= 2  # courbe seulement si au moins 2 points
    ]

    data_json = json.dumps(chart_rows, ensure_ascii=False, default=str)
    history_json = json.dumps(history_json_data, ensure_ascii=False, default=str)
    today_str = date.today().strftime("%d/%m/%Y")

    # Collect unique dates for trend data
    dates = sorted(set(r["date"] for r in chart_rows))
    all_products = list(dict.fromkeys(r["produit"] for r in chart_rows))

    html = (
        "<!DOCTYPE html>\n"
        "<html lang='fr'>\n"
        "<head>\n"
        "<meta charset='UTF-8'>\n"
        "<meta name='viewport' content='width=device-width, initial-scale=1.0'>\n"
        f"<title>HSN Whey Tracker — {today_str}</title>\n"
        "<script src='https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js'></script>\n"
        "<style>\n"
        "*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }\n"
        "body { font-family: system-ui, -apple-system, sans-serif; background: #f5f5f5;"
        " color: #222; padding: 24px; }\n"
        "h1 { font-size: 20px; font-weight: 600; margin-bottom: 4px; }\n"
        ".sub { font-size: 12px; color: #888; margin-bottom: 20px; }\n"
        ".cards { display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 24px; }\n"
        ".type-tabs { display: flex; gap: 4px; margin-bottom: 18px; border-bottom: 1px solid #e0e0e0; }\n"
        ".type-tabs button { background: none; border: none; padding: 10px 18px; font-size: 13px;"
        " color: #888; cursor: pointer; border-bottom: 2px solid transparent; margin-bottom: -1px;"
        " font-weight: 500; }\n"
        ".type-tabs button:hover { color: #333; }\n"
        ".type-tabs button.active { color: #185FA5; border-bottom-color: #378ADD; font-weight: 600; }\n"
        ".tab-hidden { display: none !important; }\n"
        ".card { background: #fff; border-radius: 10px; padding: 16px 20px; flex: 1;"
        " min-width: 150px; border: 0.5px solid #e0e0e0; }\n"
        ".card-label { font-size: 12px; color: #888; margin-bottom: 4px; }\n"
        ".card-value { font-size: 22px; font-weight: 600; }\n"
        ".card-sub { font-size: 11px; color: #888; margin-top: 2px; }\n"
        ".best { color: #0F6E56; }\n"
        ".section-title { font-size: 13px; font-weight: 600; color: #555; margin-bottom: 8px; }\n"
        ".chart-wrap { background: #fff; border-radius: 10px; padding: 20px;"
        " margin-bottom: 20px; border: 0.5px solid #e0e0e0; }\n"
        ".filters { display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 20px; }\n"
        ".filter-btn { font-size: 12px; padding: 5px 16px; border-radius: 20px;"
        " border: 1.5px solid #ccc; background: #fff; color: #555; cursor: pointer; }\n"
        ".filter-btn.active { border-color: #378ADD; background: #E6F1FB; color: #0C447C; }\n"
        "table { width: 100%; border-collapse: collapse; font-size: 12px; table-layout: fixed; }\n"
        "th { background: #f0f0f0; padding: 8px 12px; text-align: left; font-weight: 500;"
        " color: #555; border-bottom: 1px solid #e0e0e0; }\n"
        "td { padding: 7px 12px; border-bottom: 0.5px solid #eee;"
        " overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }\n"
        "tr:nth-child(even) { background: #fafafa; }\n"
        ".best-cell { color: #0F6E56; font-weight: 600; }\n"
        "a { color: #185FA5; text-decoration: none; }\n"
        ".legend { display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 10px;"
        " font-size: 12px; color: #666; }\n"
        ".legend-dot { width: 10px; height: 10px; border-radius: 2px;"
        " display: inline-block; margin-right: 4px; vertical-align: middle; }\n"
        ".table-toolbar { display: flex; gap: 12px; align-items: center; margin-bottom: 12px; }\n"
        ".search-input { padding: 6px 10px; border: 1px solid #ccc; border-radius: 6px;"
        " font-size: 12px; width: 240px; }\n"
        "th.sortable { cursor: pointer; user-select: none; }\n"
        "th.sortable:hover { background: #e8e8e8; }\n"
        "th.sortable .sort-arrow { font-size: 10px; opacity: 0.4; margin-left: 3px; }\n"
        "th.sortable.sort-asc .sort-arrow, th.sortable.sort-desc .sort-arrow { opacity: 1; color: #378ADD; }\n"
        ".deal-badge { display: inline-block; background: #FF6B35; color: #fff;"
        " font-size: 9px; font-weight: 700; padding: 1px 5px; border-radius: 3px;"
        " margin-left: 6px; vertical-align: middle; letter-spacing: 0.3px; }\n"
        ".oos-badge { display: inline-block; background: #FFF0EE; color: #B23B3B;"
        " border: 0.5px solid #F5C7C7; font-size: 9px; font-weight: 700; padding: 1px 5px;"
        " border-radius: 3px; margin-left: 6px; vertical-align: middle; }\n"
        ".trend-controls { display: flex; gap: 8px; align-items: center; flex-wrap: wrap;"
        " margin-bottom: 12px; }\n"
        ".trend-controls select { padding: 5px 8px; border: 1px solid #ccc; border-radius: 6px;"
        " font-size: 12px; max-width: 360px; }\n"
        ".filter-row { display: flex; gap: 8px; align-items: center; flex-wrap: wrap;"
        " margin-bottom: 10px; }\n"
        ".filter-row .flbl { font-size: 11px; color: #888; min-width: 110px;"
        " text-transform: uppercase; letter-spacing: 0.5px; }\n"
        ".filter-btn.sm { font-size: 11px; padding: 3px 10px; }\n"
        ".filter-btn.green.active { border-color: #1D9E75; background: #E5F6EE; color: #0F6E56; }\n"
        ".filter-btn.purple.active { border-color: #7F77DD; background: #EFEDFA; color: #4A40A8; }\n"
        ".filter-btn.amber.active { border-color: #C7892E; background: #FAF1DF; color: #7A5215; }\n"
        ".tag-chip { display: inline-block; font-size: 9px; padding: 1px 6px; border-radius: 8px;"
        " margin: 0 3px 0 0; vertical-align: middle; line-height: 14px; }\n"
        ".tag-chip.sweet { background: #FFF1F1; color: #B23B3B; border: 0.5px solid #F5C7C7; }\n"
        ".tag-chip.sweet.none { background: #ECF8F2; color: #1D7150; border-color: #C5E5D6; }\n"
        ".tag-chip.whey { background: #EEF3FB; color: #2A5C9C; border: 0.5px solid #C8D8EE; }\n"
        ".tag-chip.lbl { background: #FAF6E3; color: #7A5215; border: 0.5px solid #E8DDA8; }\n"
        ".missing-pill { display: inline-block; background: #FFF6E0; color: #92580A;"
        " border: 0.5px solid #F0DAA0; font-size: 9px; padding: 1px 6px; border-radius: 8px;"
        " margin-left: 4px; vertical-align: middle; }\n"
        ".edit-btn { background: none; border: 0; cursor: pointer; opacity: 0.4;"
        " font-size: 13px; padding: 0 4px; vertical-align: middle; }\n"
        ".edit-btn:hover { opacity: 1; }\n"
        ".note-line { color: #666; font-style: italic; font-size: 11px; margin-top: 2px;"
        " white-space: normal; }\n"
        ".export-bar { display: flex; gap: 12px; align-items: center; margin-bottom: 14px;"
        " padding: 8px 12px; background: #fff; border: 0.5px dashed #d0d0d0;"
        " border-radius: 8px; flex-wrap: wrap; }\n"
        ".export-bar button { font-size: 12px; padding: 5px 12px; border-radius: 6px;"
        " border: 1px solid #378ADD; background: #fff; color: #185FA5; cursor: pointer; }\n"
        ".export-bar button:hover { background: #E6F1FB; }\n"
        ".export-bar button:disabled { color: #aaa; border-color: #ddd; cursor: default;"
        " background: #fafafa; }\n"
        ".modal-backdrop { position: fixed; inset: 0; background: rgba(0,0,0,0.35);"
        " display: none; align-items: center; justify-content: center; z-index: 1000; }\n"
        ".modal-backdrop.open { display: flex; }\n"
        ".modal { background: #fff; border-radius: 10px; padding: 22px 24px; width: 480px;"
        " max-width: 92vw; box-shadow: 0 10px 40px rgba(0,0,0,0.2); }\n"
        ".modal h3 { font-size: 14px; margin-bottom: 4px; }\n"
        ".modal .pname { font-size: 12px; color: #888; margin-bottom: 14px; }\n"
        ".modal label { display: block; font-size: 11px; color: #666;"
        " text-transform: uppercase; letter-spacing: 0.5px; margin: 12px 0 4px; }\n"
        ".modal input, .modal textarea { width: 100%; padding: 7px 9px; font-size: 12px;"
        " border: 1px solid #ccc; border-radius: 6px; font-family: inherit; }\n"
        ".modal textarea { resize: vertical; min-height: 70px; }\n"
        ".modal .auto-block { background: #fafafa; border: 0.5px solid #eee;"
        " border-radius: 6px; padding: 8px 10px; font-size: 11px; color: #666; }\n"
        ".modal .actions { display: flex; gap: 8px; justify-content: flex-end;"
        " margin-top: 16px; }\n"
        ".modal .actions .primary { background: #378ADD; color: #fff; border-color: #378ADD; }\n"
        ".modal .actions .primary:hover { background: #2974C2; }\n"
        ".modal .actions .danger { color: #B23B3B; border-color: #E5BABA; }\n"
        ".modal .actions button { font-size: 12px; padding: 6px 14px; border-radius: 6px;"
        " border: 1px solid #ccc; background: #fff; cursor: pointer; }\n"
        "</style>\n"
        "</head>\n"
        "<body>\n"
        "<h1>HSN — Suivi des prix nutrition sportive</h1>\n"
        f"<p class='sub'>Derniere mise a jour : {today_str} &nbsp;|&nbsp; "
        f"{len(all_products)} produits &nbsp;|&nbsp; {len(dates)} jour(s) de donnees</p>\n"
        "<div class='type-tabs' id='typeTabs'></div>\n"
        "<div class='cards' id='metricCards'></div>\n"
        "<div class='filter-row' id='catRow'><span class='flbl'>Catégorie</span>"
        "<span id='categoryFilters' style='display:flex;gap:8px;flex-wrap:wrap;'></span></div>\n"
        "<div class='filter-row'><span class='flbl'>Taille</span>"
        "<span id='filters' style='display:flex;gap:8px;flex-wrap:wrap;'></span></div>\n"
        "<div class='filter-row whey-only'><span class='flbl'>Édulcorants</span>"
        "<span id='sweetFilters' style='display:flex;gap:8px;flex-wrap:wrap;'></span></div>\n"
        "<div class='filter-row whey-only'><span class='flbl'>Type protéine</span>"
        "<span id='wheyFilters' style='display:flex;gap:8px;flex-wrap:wrap;'></span></div>\n"
        "<div class='filter-row' id='typeTagRow' style='display:none;'>"
        "<span class='flbl' id='typeTagLabel'>Type</span>"
        "<span id='typeTagFilters' style='display:flex;gap:8px;flex-wrap:wrap;'></span></div>\n"
        "<div class='filter-row' id='labelFiltersRow' style='display:none;'>"
        "<span class='flbl'>Labels persos</span>"
        "<span id='labelFilters' style='display:flex;gap:8px;flex-wrap:wrap;'></span></div>\n"
        "<div class='export-bar'>\n"
        "  <span style='font-size:12px;color:#666;'>"
        "📝 <span id='editStatus'>0 modifs locales</span></span>\n"
        "  <button id='exportBtn' onclick='exportTags()'>📥 Exporter tags.json</button>\n"
        "  <button onclick='clearLocalEdits()' style='border-color:#E5BABA;color:#B23B3B;'>"
        "Annuler les modifs locales</button>\n"
        "  <span id='missingCount' style='font-size:11px;color:#92580A;margin-left:auto;'></span>\n"
        "</div>\n"
        "<div class='whey-only'>\n"
        "<div class='chart-wrap'>\n"
        "  <div class='section-title'>Prix par kilo de protéine pure (EUR/kg)</div>\n"
        "  <div class='legend' id='legendPxkgProt'></div>\n"
        "  <div style='position:relative;height:300px;'><canvas id='chartPxkgProt'></canvas></div>\n"
        "</div>\n"
        "<div class='chart-wrap'>\n"
        "  <div class='section-title'>Coût pour 30g de protéine (EUR)</div>\n"
        "  <div class='legend' id='legendCout30'></div>\n"
        "  <div style='position:relative;height:300px;'><canvas id='chartCout30'></canvas></div>\n"
        "</div>\n"
        "</div>\n"
        "<div class='chart-wrap' id='primaryChartWrap'>\n"
        "  <div class='section-title' id='primaryChartTitle'>Prix au kilo de produit (EUR/kg)</div>\n"
        "  <div class='legend' id='legendPxkg'></div>\n"
        "  <div style='position:relative;height:300px;'><canvas id='chartPxkg'></canvas></div>\n"
        "</div>\n"
        "<div class='chart-wrap whey-only'>\n"
        "  <div class='section-title'>Évolution dans le temps (EUR/kg protéine)</div>\n"
        "  <div class='trend-controls'>\n"
        "    <select id='trendAddSelect' style='flex:1;min-width:0;max-width:340px;'></select>\n"
        "    <button id='trendAddBtn' onclick='addTrendItem()'"
        " style='padding:5px 12px;border:1px solid #378ADD;background:#fff;color:#185FA5;"
        "border-radius:6px;font-size:12px;cursor:pointer;white-space:nowrap;'>+ Ajouter</button>\n"
        "    <button onclick='clearTrendItems()'"
        " style='padding:5px 12px;border:1px solid #ccc;background:#fff;color:#666;"
        "border-radius:6px;font-size:12px;cursor:pointer;'>Tout effacer</button>\n"
        "  </div>\n"
        "  <div id='trendSelected' style='display:flex;gap:6px;flex-wrap:wrap;"
        "min-height:24px;margin-top:8px;margin-bottom:4px;'></div>\n"
        "  <div style='position:relative;height:280px;'><canvas id='chartTrend'></canvas></div>\n"
        "</div>\n"
        "<div class='chart-wrap'>\n"
        "  <div class='section-title'>Données complètes</div>\n"
        "  <div class='table-toolbar'>\n"
        "    <input id='searchInput' class='search-input' type='text' placeholder='Rechercher un produit...'>\n"
        "    <span id='rowCount' style='font-size:11px;color:#888'></span>\n"
        "  </div>\n"
        "  <table id='detailTable'><thead id='tableHead'></thead><tbody id='tableBody'></tbody></table>\n"
        "</div>\n"
        # ── Modal d'édition tags / note ─────────────────────────────────────
        "<div class='modal-backdrop' id='editModal' onclick='if(event.target.id===\"editModal\")closeEditModal()'>\n"
        "  <div class='modal'>\n"
        "    <h3>Tags &amp; note du produit</h3>\n"
        "    <div class='pname' id='modalProduit'></div>\n"
        "    <div class='auto-block' id='modalAuto'></div>\n"
        "    <label>Labels persos (séparés par virgule)</label>\n"
        "    <input id='modalLabels' type='text' placeholder='ex: favori, à tester, best-deal-pro'>\n"
        "    <label>Note libre</label>\n"
        "    <textarea id='modalNote' placeholder='Goût, observations, …'></textarea>\n"
        "    <div class='actions'>\n"
        "      <button class='danger' onclick='resetProductEdit()'>Réinitialiser</button>\n"
        "      <button onclick='closeEditModal()'>Annuler</button>\n"
        "      <button class='primary' onclick='saveEdit()'>Enregistrer</button>\n"
        "    </div>\n"
        "  </div>\n"
        "</div>\n"
        f"<script>\nconst RAW = {data_json};\n"
        f"const HISTORY = {history_json};\n"
        f"const SWEETENER_LABELS = {json.dumps(SWEETENER_LABELS, ensure_ascii=False)};\n"
        f"const WHEY_TYPE_LABELS = {json.dumps(WHEY_TYPE_LABELS, ensure_ascii=False)};\n"
        f"const OMEGA3_TYPE_LABELS = {json.dumps(OMEGA3_TYPE_LABELS, ensure_ascii=False)};\n"
        f"const CREATINE_TYPE_LABELS = {json.dumps(CREATINE_TYPE_LABELS, ensure_ascii=False)};\n"
        "const TYPE_TAB_LABEL = { whey:'Whey & protéines', omega3:'Oméga-3', creatine:'Créatine' };\n"
        "// Colonnes du tableau par type de produit\n"
        "const TAB_COLS = {\n"
        "  whey: [\n"
        "    {k:'produit',l:'Produit',w:'26%',wide:true},\n"
        "    {k:'taille',l:'Taille',w:'7%'},\n"
        "    {k:'prix',l:'Prix',w:'7%',num:true,d:2},\n"
        "    {k:'prot',l:'%Prot',w:'6%',num:true,fmt:'pct'},\n"
        "    {k:'pxkgProt',l:'EUR/kg prot',w:'11%',num:true,d:2,best:true},\n"
        "    {k:'cout30',l:'EUR/30g prot',w:'11%',num:true,d:3,best:true},\n"
        "    {k:'cout3leu',l:'EUR/3g leu',w:'11%',num:true,d:3,best:true,title:'Coût pour 3g de leucine (seuil anabolique)'},\n"
        "    {k:'pxkg',l:'EUR/kg',w:'8%',num:true,d:2},\n"
        "    {k:'cout',l:'EUR/portion',w:'8%',num:true,d:2},\n"
        "    {k:'date',l:'Date',w:'5%',num:true,fmt:'date'},\n"
        "  ],\n"
        "  omega3: [\n"
        "    {k:'produit',l:'Produit',w:'30%',wide:true},\n"
        "    {k:'taille',l:'Format',w:'10%'},\n"
        "    {k:'prix',l:'Prix',w:'8%',num:true,d:2},\n"
        "    {k:'epa',l:'EPA mg/cap',w:'10%',num:true,fmt:'int'},\n"
        "    {k:'dha',l:'DHA mg/cap',w:'10%',num:true,fmt:'int'},\n"
        "    {k:'coutGOmega',l:'EUR/g EPA+DHA',w:'14%',num:true,d:3,best:true},\n"
        "    {k:'pxkg',l:'EUR/kg',w:'10%',num:true,d:2},\n"
        "    {k:'date',l:'Date',w:'8%',num:true,fmt:'date'},\n"
        "  ],\n"
        "  creatine: [\n"
        "    {k:'produit',l:'Produit',w:'40%',wide:true},\n"
        "    {k:'taille',l:'Taille',w:'10%'},\n"
        "    {k:'prix',l:'Prix',w:'10%',num:true,d:2},\n"
        "    {k:'creaDose',l:'g/dose',w:'10%',num:true,fmt:'int'},\n"
        "    {k:'coutKgCrea',l:'EUR/kg créatine',w:'15%',num:true,d:2,best:true},\n"
        "    {k:'pxkg',l:'EUR/kg',w:'10%',num:true,d:2},\n"
        "    {k:'date',l:'Date',w:'10%',num:true,fmt:'date'},\n"
        "  ],\n"
        "};\n"
        "// Métrique principale par tab : sort key par défaut + chart key + label\n"
        "const TAB_PRIMARY = {\n"
        "  whey:     {sort:'pxkgProt', key:'pxkg',       label:'Prix au kilo de produit (EUR/kg)'},\n"
        "  omega3:   {sort:'coutGOmega', key:'coutGOmega', label:'Coût pour 1g d\\'EPA+DHA (EUR)'},\n"
        "  creatine: {sort:'coutKgCrea', key:'coutKgCrea', label:'Coût au kilo de créatine (EUR/kg)'},\n"
        "};\n"
        "const COLORS = {'500g':'#378ADD','750g':'#1D9E75','2Kg':'#7F77DD','Unique':'#D85A30'};\n"
        "const TREND_COLORS = ['#378ADD','#1D9E75','#E84040','#F5A623','#7F77DD','#00BCD4','#FF6B35','#8BC34A'];\n"
        "const CAT_ORDER = ['Whey','Aliments enrichis','Autres'];\n"
        "const CATEGORIES = CAT_ORDER.filter(c=>RAW.some(r=>r.categorie===c)).concat([...new Set(RAW.map(r=>r.categorie).filter(c=>c && !CAT_ORDER.includes(c)))]);\n"
        "const PRODUCTS = [...new Set(RAW.map(r=>r.produit))];\n"
        "const SIZES = [...new Set(RAW.map(r=>r.taille))];\n"
        "const SWEETENERS = Object.keys(SWEETENER_LABELS).filter(k=>RAW.some(r=>(r.sweeteners||[]).includes(k)));\n"
        "const WHEY_TYPES = Object.keys(WHEY_TYPE_LABELS).filter(k=>RAW.some(r=>(r.wheyTypes||[]).includes(k)));\n"
        "const STORAGE_KEY = 'hsn_tracker_local_tags_v1';\n"
        "// Tags chargés depuis tags.json (server-side) + localStorage (édits locaux)\n"
        "const SERVER_TAGS = {};\n"
        "RAW.forEach(r=>{ if(r.labels && r.labels.length || (r.note||'').length){ SERVER_TAGS[r.produit] = {labels:[...r.labels], note:r.note}; }});\n"
        "function loadLocalTags(){ try { return JSON.parse(localStorage.getItem(STORAGE_KEY)||'{}'); } catch(e){ return {}; } }\n"
        "function saveLocalTags(t){ localStorage.setItem(STORAGE_KEY, JSON.stringify(t)); }\n"
        "let LOCAL_TAGS = loadLocalTags();\n"
        "function effectiveTags(produit){\n"
        "  const local = LOCAL_TAGS[produit];\n"
        "  if(local) return local;\n"
        "  return SERVER_TAGS[produit] || {labels:[], note:''};\n"
        "}\n"
        "let currentSize = 'all';\n"
        "let currentCategory = 'Whey';\n"
        "let currentTab = 'whey';\n"
        "let selectedSweeteners = new Set();\n"
        "let selectedWheyTypes = new Set();\n"
        "let selectedLabels = new Set();\n"
        "let selectedTypeTags = new Set();\n"
        "let swLogic = 'OR', wtLogic = 'OR', lblLogic = 'OR', ttLogic = 'OR';\n"
        "let searchQuery = '';\n"
        "let sortKey = 'pxkgProt', sortAsc = true;\n"
        "let chartPxkgProt, chartCout30, chartPxkg, chartTrend;\n"
        "let selectedTrendIndices = [];\n"
        "let modalCurrent = null;\n"
        "function fmt(v,d=2,suf=' EUR'){return v!=null?Number(v).toFixed(d)+suf:'—';}\n"
        "function matchesGroup(values, selected, mode){\n"
        "  if(!selected.size) return true;\n"
        "  if(mode==='AND'){ for(const t of selected) if(!values.includes(t)) return false; return true; }\n"
        "  for(const t of selected) if(values.includes(t)) return true;\n"
        "  return false;\n"
        "}\n"
        "function getFiltered(){\n"
        "  let f = RAW.filter(r => (r.type||'whey') === currentTab);\n"
        "  if(currentTab==='whey' && currentCategory!=='all'){f=f.filter(r=>r.categorie===currentCategory);}\n"
        "  if(currentSize!=='all'){f=f.filter(r=>r.taille===currentSize);}\n"
        "  if(currentTab==='whey'){\n"
        "    f = f.filter(r => matchesGroup(r.sweeteners||[], selectedSweeteners, swLogic));\n"
        "    f = f.filter(r => matchesGroup(r.wheyTypes||[], selectedWheyTypes, wtLogic));\n"
        "  } else {\n"
        "    f = f.filter(r => matchesGroup(r.typeTags||[], selectedTypeTags, ttLogic));\n"
        "  }\n"
        "  f = f.filter(r => matchesGroup(effectiveTags(r.produit).labels||[], selectedLabels, lblLogic));\n"
        "  if(searchQuery){const q=searchQuery.toLowerCase();f=f.filter(r=>r.produit.toLowerCase().includes(q)||(effectiveTags(r.produit).note||'').toLowerCase().includes(q));}\n"
        "  return f;\n"
        "}\n"
        "function shortName(p){return p.length>32?p.slice(0,30)+'…':p;}\n"
        "function getFilteredProducts(){return [...new Set(getFiltered().map(r=>r.produit))];}\n"
        "function buildDatasets(key){\n"
        "  const products=getFilteredProducts();\n"
        "  const sizes=currentSize==='all'?SIZES:[currentSize];\n"
        "  return sizes.map(sz=>({\n"
        "    label:sz,\n"
        "    data:products.map(pr=>{const r=RAW.find(x=>x.produit===pr&&x.taille===sz&&(currentCategory==='all'||x.categorie===currentCategory));return r?r[key]:null;}),\n"
        "    backgroundColor:COLORS[sz]||'#888',borderRadius:4,borderSkipped:false\n"
        "  }));\n"
        "}\n"
        "const COPTS=(unit=' EUR',d=2)=>({\n"
        "  responsive:true,maintainAspectRatio:false,\n"
        "  plugins:{legend:{display:false},tooltip:{callbacks:{label:ctx=>`${ctx.dataset.label}: ${ctx.parsed.y!=null?Number(ctx.parsed.y).toFixed(d)+unit:'N/A'}`}}},\n"
        "  scales:{x:{grid:{display:false},ticks:{font:{size:11},maxRotation:35}},\n"
        "          y:{grid:{color:'rgba(0,0,0,0.06)'},ticks:{font:{size:11},callback:v=>Number(v).toFixed(d)+unit},beginAtZero:false}}\n"
        "});\n"
        "function buildLegend(id){\n"
        "  const sizes=currentSize==='all'?SIZES:[currentSize];\n"
        "  document.getElementById(id).innerHTML=sizes.map(sz=>`<span><span class='legend-dot' style='background:${COLORS[sz]||\"#888\"}'></span>${sz}</span>`).join('');\n"
        "}\n"
        "function initCharts(){\n"
        "  const labels=getFilteredProducts().map(shortName);\n"
        "  chartPxkgProt=new Chart(document.getElementById('chartPxkgProt'),{type:'bar',data:{labels,datasets:buildDatasets('pxkgProt')},options:COPTS()});\n"
        "  chartCout30=new Chart(document.getElementById('chartCout30'),{type:'bar',data:{labels,datasets:buildDatasets('cout30')},options:COPTS(' EUR',3)});\n"
        "  chartPxkg=new Chart(document.getElementById('chartPxkg'),{type:'bar',data:{labels,datasets:buildDatasets('pxkg')},options:COPTS()});\n"
        "  buildLegend('legendPxkgProt');buildLegend('legendCout30');buildLegend('legendPxkg');\n"
        "}\n"
        "function updateCharts(){\n"
        "  const labels=getFilteredProducts().map(shortName);\n"
        "  if(currentTab==='whey'){\n"
        "    [[chartPxkgProt,'pxkgProt'],[chartCout30,'cout30'],[chartPxkg,'pxkg']].forEach(([c,k])=>{\n"
        "      c.data.labels=labels; c.data.datasets=buildDatasets(k); c.update();\n"
        "    });\n"
        "    document.getElementById('primaryChartTitle').textContent='Prix au kilo de produit (EUR/kg)';\n"
        "    buildLegend('legendPxkgProt');buildLegend('legendCout30');buildLegend('legendPxkg');\n"
        "  } else {\n"
        "    const prim=TAB_PRIMARY[currentTab];\n"
        "    document.getElementById('primaryChartTitle').textContent=prim.label;\n"
        "    chartPxkg.data.labels=labels;\n"
        "    chartPxkg.data.datasets=buildDatasets(prim.key);\n"
        "    chartPxkg.update();\n"
        "    buildLegend('legendPxkg');\n"
        "  }\n"
        "}\n"
        "function buildCategoryFilters(){\n"
        "  const cats=['all',...CATEGORIES];\n"
        "  document.getElementById('categoryFilters').innerHTML=cats.map(c=>`<button class='filter-btn ${c===currentCategory?\"active\":\"\"}' onclick='filterCategory(\"${c}\")'>${c==='all'?'Toutes catégories':c}</button>`).join('');\n"
        "}\n"
        "function filterCategory(c){currentCategory=c;buildCategoryFilters();updateCharts();updateTable();"
        "selectedTrendIndices=[];renderTrendChips();buildTrendSelect();buildTrendChart();}\n"
        "function compareValues(a,b,k){\n"
        "  const av=a[k], bv=b[k];\n"
        "  if(av==null && bv==null) return 0;\n"
        "  if(av==null) return 1;\n"
        "  if(bv==null) return -1;\n"
        "  if(typeof av === 'number' && typeof bv === 'number') return av-bv;\n"
        "  return String(av).localeCompare(String(bv));\n"
        "}\n"
        "function escapeHtml(s){return String(s||'').replace(/[&<>\"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',\"'\":'&#39;'}[c]));}\n"
        "function typeTagLabel(t){\n"
        "  return OMEGA3_TYPE_LABELS[t] || CREATINE_TYPE_LABELS[t] || t;\n"
        "}\n"
        "function renderChips(r){\n"
        "  const parts=[];\n"
        "  for(const t of (r.sweeteners||[])){\n"
        "    const cls=t==='sans_edulcorant'?'sweet none':'sweet';\n"
        "    parts.push(`<span class='tag-chip ${cls}'>${SWEETENER_LABELS[t]||t}</span>`);\n"
        "  }\n"
        "  for(const t of (r.wheyTypes||[])){\n"
        "    parts.push(`<span class='tag-chip whey'>${WHEY_TYPE_LABELS[t]||t}</span>`);\n"
        "  }\n"
        "  for(const t of (r.typeTags||[])){\n"
        "    parts.push(`<span class='tag-chip whey'>${escapeHtml(typeTagLabel(t))}</span>`);\n"
        "  }\n"
        "  const eff=effectiveTags(r.produit);\n"
        "  for(const lb of (eff.labels||[])){\n"
        "    parts.push(`<span class='tag-chip lbl'>${escapeHtml(lb)}</span>`);\n"
        "  }\n"
        "  return parts.join('');\n"
        "}\n"
        "function isMissingInfo(r){\n"
        "  // Manque de classification : pas d'ingrédients OU ni édulcorant ni type protéine détecté\n"
        "  if(!r.hasIngredients) return true;\n"
        "  return (!(r.sweeteners||[]).length) && (!(r.wheyTypes||[]).length);\n"
        "}\n"
        "function fmtCell(col, v){\n"
        "  if(v==null||v==='') return '—';\n"
        "  if(col.fmt==='pct') return Number(v).toFixed(0)+'%';\n"
        "  if(col.fmt==='int') return Number(v).toFixed(0);\n"
        "  if(col.fmt==='date') return v;\n"
        "  if(col.num) return Number(v).toFixed(col.d!=null?col.d:2)+' EUR';\n"
        "  return String(v);\n"
        "}\n"
        "function renderTableHead(){\n"
        "  const cols=TAB_COLS[currentTab]||TAB_COLS.whey;\n"
        "  const html='<tr>'+cols.map(c=>{\n"
        "    const align=c.num?';text-align:right':'';\n"
        "    const t=c.title?` title='${escapeHtml(c.title)}'`:'';\n"
        "    return `<th class='sortable' data-key='${c.k}'${t} style='width:${c.w||\"auto\"}${align}'>${escapeHtml(c.l)}<span class='sort-arrow'>↕</span></th>`;\n"
        "  }).join('')+'</tr>';\n"
        "  document.getElementById('tableHead').innerHTML=html;\n"
        "  setupSort();\n"
        "  updateSortHeaders();\n"
        "}\n"
        "function updateTable(){\n"
        "  const cols=TAB_COLS[currentTab]||TAB_COLS.whey;\n"
        "  let f=getFiltered();\n"
        "  f=[...f].sort((a,b)=>{const c=compareValues(a,b,sortKey);return sortAsc?c:-c;});\n"
        "  // Calcule les meilleures valeurs (min) pour les colonnes avec best:true\n"
        "  const bestVals={};\n"
        "  for(const c of cols){ if(c.best){\n"
        "    const vals=f.filter(r=>r[c.k]!=null).map(r=>r[c.k]);\n"
        "    if(vals.length) bestVals[c.k]=Math.min(...vals);\n"
        "  }}\n"
        "  document.getElementById('rowCount').textContent=`${f.length} ligne(s)`;\n"
        "  if(f.length===0){\n"
        "    const colspan = cols.length;\n"
        "    const tabRows = RAW.filter(r=>(r.type||'whey')===currentTab);\n"
        "    const msg = tabRows.length===0\n"
        "      ? `Aucune donnée pour <b>${TYPE_TAB_LABEL[currentTab]}</b> dans l'Excel.<br>Lance <code>python hsn_tracker.py</code> pour scraper ${currentTab==='omega3'?'les oméga-3':currentTab==='creatine'?'les créatines':'les whey'}.`\n"
        "      : `Aucun produit ne correspond aux filtres actuels.`;\n"
        "    document.getElementById('tableBody').innerHTML = `<tr><td colspan='${colspan}' style='padding:30px;text-align:center;color:#888;font-size:13px'>${msg}</td></tr>`;\n"
        "    return;\n"
        "  }\n"
        "  document.getElementById('tableBody').innerHTML=f.map(r=>{\n"
        "    const dealBadge = r.isDeal ? `<span class='deal-badge' title='Sous la moyenne historique de plus de 5%'>🔥 DEAL</span>` : '';\n"
        "    const missing = (currentTab==='whey' && isMissingInfo(r)) ? `<span class='missing-pill' title='Pas (ou peu) d&apos;info détectée — clic ✏️ pour annoter'>❓ à classer</span>` : '';\n"
        "    const oosBadge = r.en_stock === false ? `<span class='oos-badge' title='Rupture de stock détectée au dernier scraping'>⚠️ Rupture</span>` : '';\n"
        "    const eff = effectiveTags(r.produit);\n"
        "    const noteLine = eff.note ? `<div class='note-line'>📝 ${escapeHtml(eff.note)}</div>` : '';\n"
        "    const editBtn = `<button class='edit-btn' title='Éditer tags &amp; note' onclick='openEditModal(${escapeHtml(JSON.stringify(r.produit))})'>✏️</button>`;\n"
        "    const cells = cols.map(c=>{\n"
        "      if(c.k==='produit'){\n"
        "        return `<td title='${escapeHtml(r.produit)}' style='white-space:normal;overflow:visible;'><a href='${r.url}' target='_blank'>${escapeHtml(r.produit)}</a>${editBtn}${dealBadge}${missing}${oosBadge}<div style='margin-top:3px;white-space:normal;'>${renderChips(r)}</div>${noteLine}</td>`;\n"
        "      }\n"
        "      const align = c.num ? 'text-align:right' : '';\n"
        "      const isBest = c.best && bestVals[c.k]!=null && r[c.k]===bestVals[c.k];\n"
        "      const cls = isBest ? 'best-cell' : '';\n"
        "      const dateStyle = c.fmt==='date' ? ';font-size:11px;color:#888' : '';\n"
        "      return `<td class='${cls}' style='${align}${dateStyle}'>${fmtCell(c, r[c.k])}</td>`;\n"
        "    }).join('');\n"
        "    return `<tr>${cells}</tr>`;\n"
        "  }).join('');\n"
        "  // Indicateur global \"à classer\" (whey uniquement)\n"
        "  if(currentTab==='whey'){\n"
        "    const missCount = RAW.filter(r=>(r.type||'whey')==='whey').filter(isMissingInfo).map(r=>r.produit);\n"
        "    const uniqMiss = [...new Set(missCount)].length;\n"
        "    document.getElementById('missingCount').textContent = uniqMiss>0 ? `❓ ${uniqMiss} produit(s) à classer` : '';\n"
        "  } else {\n"
        "    document.getElementById('missingCount').textContent = '';\n"
        "  }\n"
        "}\n"
        "function updateSortHeaders(){\n"
        "  document.querySelectorAll('th.sortable').forEach(th=>{\n"
        "    th.classList.remove('sort-asc','sort-desc');\n"
        "    if(th.dataset.key===sortKey){th.classList.add(sortAsc?'sort-asc':'sort-desc');\n"
        "      th.querySelector('.sort-arrow').textContent=sortAsc?'↑':'↓';}\n"
        "    else{th.querySelector('.sort-arrow').textContent='↕';}\n"
        "  });\n"
        "}\n"
        "function setupSort(){\n"
        "  document.querySelectorAll('th.sortable').forEach(th=>{\n"
        "    th.addEventListener('click',()=>{\n"
        "      const k=th.dataset.key;\n"
        "      if(sortKey===k){sortAsc=!sortAsc;}else{sortKey=k;sortAsc=true;}\n"
        "      updateSortHeaders();updateTable();\n"
        "    });\n"
        "  });\n"
        "}\n"
        "function setupSearch(){\n"
        "  document.getElementById('searchInput').addEventListener('input',e=>{\n"
        "    searchQuery=e.target.value.trim();updateTable();\n"
        "  });\n"
        "}\n"
        "function buildTrendSelect(){\n"
        "  const sel=document.getElementById('trendAddSelect');\n"
        "  const filtered=HISTORY.map((h,i)=>({h,i})).filter(x=>currentCategory==='all'||x.h.categorie===currentCategory);\n"
        "  sel.innerHTML=filtered.map(({h,i})=>`<option value='${i}'>${h.produit} — ${h.taille} (${h.points.length} pts)</option>`).join('');\n"
        "  if(filtered.length===0){sel.innerHTML='<option>Pas assez de données (besoin de 2+ jours)</option>';}\n"
        "}\n"
        "function addTrendItem(){\n"
        "  const idx=parseInt(document.getElementById('trendAddSelect').value);\n"
        "  if(isNaN(idx)||selectedTrendIndices.length>=8) return;\n"
        "  if(!selectedTrendIndices.includes(idx)){\n"
        "    selectedTrendIndices.push(idx);\n"
        "    renderTrendChips();\n"
        "    buildTrendChart();\n"
        "  }\n"
        "}\n"
        "function removeTrendItem(idx){\n"
        "  selectedTrendIndices=selectedTrendIndices.filter(i=>i!==idx);\n"
        "  renderTrendChips();\n"
        "  buildTrendChart();\n"
        "}\n"
        "function clearTrendItems(){\n"
        "  selectedTrendIndices=[];\n"
        "  renderTrendChips();\n"
        "  buildTrendChart();\n"
        "}\n"
        "function renderTrendChips(){\n"
        "  const container=document.getElementById('trendSelected');\n"
        "  if(!selectedTrendIndices.length){\n"
        "    container.innerHTML=`<span style='font-size:11px;color:#aaa;align-self:center;'>"
        "Sélectionne un produit et clique « + Ajouter » pour tracer son évolution.</span>`;\n"
        "    document.getElementById('trendAddBtn').disabled=false;\n"
        "    return;\n"
        "  }\n"
        "  container.innerHTML=selectedTrendIndices.map((idx,i)=>{\n"
        "    const h=HISTORY[idx]; if(!h) return '';\n"
        "    const c=TREND_COLORS[i%TREND_COLORS.length];\n"
        "    return `<span style='display:inline-flex;align-items:center;gap:4px;background:#fff;"
        "border:1.5px solid ${c};color:#333;border-radius:12px;padding:2px 8px 2px 6px;"
        "font-size:11px;'>`\n"
        "      + `<span style='width:8px;height:8px;border-radius:50%;background:${c};"
        "display:inline-block;flex-shrink:0;'></span>`\n"
        "      + `${escapeHtml(h.produit.slice(0,28))} — ${escapeHtml(h.taille)}`\n"
        "      + `<button onclick='removeTrendItem(${idx})' style='background:none;border:0;"
        "cursor:pointer;color:#999;font-size:14px;padding:0 0 0 4px;line-height:1;'>×</button>`\n"
        "      + `</span>`;\n"
        "  }).join('');\n"
        "  document.getElementById('trendAddBtn').disabled=selectedTrendIndices.length>=8;\n"
        "}\n"
        "function buildTrendChart(){\n"
        "  if(chartTrend){chartTrend.destroy(); chartTrend=null;}\n"
        "  if(!selectedTrendIndices.length){\n"
        "    chartTrend=new Chart(document.getElementById('chartTrend'),{\n"
        "      type:'line',data:{labels:[],datasets:[]},\n"
        "      options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}}}\n"
        "    });\n"
        "    return;\n"
        "  }\n"
        "  const allDates=[...new Set(selectedTrendIndices.flatMap(idx=>(HISTORY[idx]||{points:[]}).points.map(p=>p.date)))].sort();\n"
        "  const datasets=selectedTrendIndices.map((idx,i)=>{\n"
        "    const h=HISTORY[idx]; if(!h) return null;\n"
        "    const color=TREND_COLORS[i%TREND_COLORS.length];\n"
        "    const dm={}; h.points.forEach(p=>{dm[p.date]=p.pxkgProt;});\n"
        "    return{label:`${h.produit.slice(0,24)} — ${h.taille}`,\n"
        "      data:allDates.map(d=>dm[d]??null),\n"
        "      borderColor:color,backgroundColor:color+'22',\n"
        "      tension:0.2,pointRadius:4,fill:false,spanGaps:true};\n"
        "  }).filter(Boolean);\n"
        "  chartTrend=new Chart(document.getElementById('chartTrend'),{\n"
        "    type:'line',\n"
        "    data:{labels:allDates,datasets},\n"
        "    options:{responsive:true,maintainAspectRatio:false,\n"
        "      plugins:{legend:{display:datasets.length>1,position:'top',labels:{font:{size:11},"
        "boxWidth:12}},\n"
        "        tooltip:{callbacks:{label:ctx=>`${ctx.dataset.label}: ${ctx.parsed.y!=null?Number(ctx.parsed.y).toFixed(2)+' EUR':'N/A'}`}}},\n"
        "      scales:{x:{grid:{display:false},ticks:{font:{size:11},maxRotation:35}},\n"
        "        y:{grid:{color:'rgba(0,0,0,0.06)'},ticks:{font:{size:11},"
        "callback:v=>Number(v).toFixed(2)+' EUR'},beginAtZero:false}}}\n"
        "  });\n"
        "}\n"
        "function setupTrend(){}\n"
        "function bestOf(rows, key){ const w=rows.filter(r=>r[key]!=null); return w.length?w.reduce((a,b)=>a[key]<b[key]?a:b):null; }\n"
        "function cardHtml(label,value,sub,best){\n"
        "  const cls = best?'card-value best':'card-value';\n"
        "  const subHtml = sub?`<div class='card-sub'>${escapeHtml(sub)}</div>`:'';\n"
        "  return `<div class='card'><div class='card-label'>${escapeHtml(label)}</div><div class='${cls}'>${value}</div>${subHtml}</div>`;\n"
        "}\n"
        "function buildMetrics(){\n"
        "  const tabRows = RAW.filter(r=>(r.type||'whey')===currentTab);\n"
        "  const prods=new Set(tabRows.map(r=>r.produit)).size;\n"
        "  const cards=[];\n"
        "  cards.push(`<div class='card'><div class='card-label'>${escapeHtml(TYPE_TAB_LABEL[currentTab]||'Produits')}</div><div class='card-value'>${prods}</div></div>`);\n"
        "  if(currentTab==='whey'){\n"
        "    const bp=bestOf(tabRows,'pxkgProt'), bc=bestOf(tabRows,'cout30'), bl=bestOf(tabRows,'cout3leu'), bx=bestOf(tabRows,'pxkg');\n"
        "    if(bp) cards.push(cardHtml('Meilleur EUR/kg protéine', bp.pxkgProt.toFixed(2)+' EUR', `${bp.produit.slice(0,28)} — ${bp.taille}`, true));\n"
        "    if(bc) cards.push(cardHtml('Meilleur 30g protéine', bc.cout30.toFixed(3)+' EUR', `${bc.produit.slice(0,28)} — ${bc.taille}`, true));\n"
        "    if(bl) cards.push(cardHtml('Meilleur 3g leucine', bl.cout3leu.toFixed(3)+' EUR', `${bl.produit.slice(0,28)} — ${bl.taille}`, true));\n"
        "    if(bx) cards.push(cardHtml('Meilleur EUR/kg produit', bx.pxkg.toFixed(2)+' EUR', `${bx.produit.slice(0,28)} — ${bx.taille}`));\n"
        "  } else if(currentTab==='omega3'){\n"
        "    const bo=bestOf(tabRows,'coutGOmega'), bx=bestOf(tabRows,'pxkg');\n"
        "    if(bo) cards.push(cardHtml('Meilleur EUR/g EPA+DHA', bo.coutGOmega.toFixed(3)+' EUR', `${bo.produit.slice(0,28)} — ${bo.taille}`, true));\n"
        "    else cards.push(cardHtml('EUR/g EPA+DHA', '— données manquantes —', 'Lance le scraper pour collecter EPA/DHA'));\n"
        "    if(bx) cards.push(cardHtml('Meilleur EUR/kg', bx.pxkg.toFixed(2)+' EUR', `${bx.produit.slice(0,28)} — ${bx.taille}`));\n"
        "  } else if(currentTab==='creatine'){\n"
        "    const bc=bestOf(tabRows,'coutKgCrea'), bx=bestOf(tabRows,'pxkg');\n"
        "    if(bc) cards.push(cardHtml('Meilleur EUR/kg créatine', bc.coutKgCrea.toFixed(2)+' EUR', `${bc.produit.slice(0,28)} — ${bc.taille}`, true));\n"
        "    if(bx) cards.push(cardHtml('Meilleur EUR/kg', bx.pxkg.toFixed(2)+' EUR', `${bx.produit.slice(0,28)} — ${bx.taille}`));\n"
        "  }\n"
        f"  cards.push(`<div class='card'><div class='card-label'>Mise à jour</div><div class='card-value' style='font-size:16px'>{today_str}</div></div>`);\n"
        "  document.getElementById('metricCards').innerHTML=cards.join('');\n"
        "}\n"
        # ── Type tabs (Whey / Oméga-3 / Créatine) ───────────────────────────
        "function tabCount(t){ return RAW.filter(r=>(r.type||'whey')===t).length; }\n"
        "function buildTypeTabs(){\n"
        "  // Toujours afficher les 3 onglets ; un onglet vide est grisé + compteur (0)\n"
        "  const tabs = ['whey','omega3','creatine'];\n"
        "  document.getElementById('typeTabs').innerHTML = tabs.map(t=>{\n"
        "    const n = tabCount(t);\n"
        "    const dim = n===0 ? ' style=\"opacity:0.45\"' : '';\n"
        "    return `<button class='${t===currentTab?\"active\":\"\"}'${dim} onclick='setTab(\"${t}\")'>`\n"
        "         + `${TYPE_TAB_LABEL[t]||t} <span style='color:#aaa;font-size:11px'>(${n})</span></button>`;\n"
        "  }).join('');\n"
        "}\n"
        "function setTab(t){\n"
        "  if(t===currentTab) return;\n"
        "  currentTab=t;\n"
        "  // Reset des filtres spécifiques à un tab\n"
        "  selectedSweeteners.clear(); selectedWheyTypes.clear(); selectedTypeTags.clear();\n"
        "  currentSize='all'; currentCategory = (currentTab==='whey') ? 'Whey' : 'all';\n"
        "  // Sort key par défaut du tab\n"
        "  sortKey = (TAB_PRIMARY[currentTab]||TAB_PRIMARY.whey).sort;\n"
        "  sortAsc = true;\n"
        "  applyTabUI();\n"
        "  buildTypeTabs(); buildCategoryFilters(); buildFilters();\n"
        "  buildSweetFilters(); buildWheyFilters(); buildTypeTagFilters(); buildLabelFilters();\n"
        "  renderTableHead();\n"
        "  buildMetrics();\n"
        "  updateCharts();\n"
        "  updateTable();\n"
        "  selectedTrendIndices=[]; renderTrendChips(); buildTrendSelect(); buildTrendChart();\n"
        "}\n"
        "function applyTabUI(){\n"
        "  // .whey-only éléments visibles uniquement sur le tab whey\n"
        "  document.querySelectorAll('.whey-only').forEach(el=>{\n"
        "    el.classList.toggle('tab-hidden', currentTab!=='whey');\n"
        "  });\n"
        "  // catRow visible uniquement sur whey (Whey/Aliments enrichis/Autres)\n"
        "  document.getElementById('catRow').classList.toggle('tab-hidden', currentTab!=='whey');\n"
        "  // typeTagRow visible uniquement sur omega3/creatine\n"
        "  const typeTagRow = document.getElementById('typeTagRow');\n"
        "  typeTagRow.style.display = currentTab==='whey' ? 'none' : 'flex';\n"
        "  if(currentTab==='omega3') document.getElementById('typeTagLabel').textContent='Forme';\n"
        "  if(currentTab==='creatine') document.getElementById('typeTagLabel').textContent='Type';\n"
        "}\n"
        "function buildTypeTagFilters(){\n"
        "  if(currentTab==='whey'){ document.getElementById('typeTagFilters').innerHTML=''; return; }\n"
        "  const labelMap = currentTab==='omega3' ? OMEGA3_TYPE_LABELS : CREATINE_TYPE_LABELS;\n"
        "  const present = Object.keys(labelMap).filter(k=>RAW.some(r=>(r.type||'')===currentTab && (r.typeTags||[]).includes(k)));\n"
        "  const chips = present.map(t=>`<button class='filter-btn sm purple ${selectedTypeTags.has(t)?\"active\":\"\"}' onclick='toggleTypeTag(\"${t}\")'>${escapeHtml(labelMap[t])}</button>`).join('');\n"
        "  const lc = `<button class='filter-btn sm' style='border-color:#bbb;background:#f8f8f8;color:#666' onclick='toggleLogic(\"tt\")'>${ttLogic==='AND'?'ET':'OU'}</button>`;\n"
        "  document.getElementById('typeTagFilters').innerHTML = present.length ? (lc + chips) : `<span style='font-size:11px;color:#aaa'>aucun détecté — lance le scraper</span>`;\n"
        "}\n"
        "function toggleTypeTag(t){toggleSet(selectedTypeTags,t);buildTypeTagFilters();updateTable();updateCharts();}\n"
        "function buildFilters(){\n"
        "  document.getElementById('filters').innerHTML=['all',...SIZES].map(sz=>`\n"
        "    <button class='filter-btn sm ${sz===currentSize?\"active\":\"\"}'"
        " onclick='filterSize(\"${sz}\")'>${sz==='all'?'Toutes':sz}</button>`).join('');\n"
        "}\n"
        "function filterSize(sz){currentSize=sz;buildFilters();updateCharts();updateTable();buildTrendSelect();}\n"
        # ── Sweetener / whey-type / labels filters ─────────────────────────
        "function toggleSet(set,val){ if(set.has(val)) set.delete(val); else set.add(val); }\n"
        "function logicChip(group, mode){\n"
        "  return `<button class='filter-btn sm' style='border-color:#bbb;background:#f8f8f8;color:#666' "
        "title='Bascule ET / OU' onclick='toggleLogic(\"${group}\")'>"
        "${mode==='AND'?'ET':'OU'}</button>`;\n"
        "}\n"
        "function toggleLogic(group){\n"
        "  if(group==='sw') swLogic = swLogic==='OR'?'AND':'OR';\n"
        "  if(group==='wt') wtLogic = wtLogic==='OR'?'AND':'OR';\n"
        "  if(group==='lbl') lblLogic = lblLogic==='OR'?'AND':'OR';\n"
        "  if(group==='tt') ttLogic = ttLogic==='OR'?'AND':'OR';\n"
        "  buildSweetFilters(); buildWheyFilters(); buildTypeTagFilters(); buildLabelFilters();\n"
        "  updateTable(); updateCharts();\n"
        "}\n"
        "function buildSweetFilters(){\n"
        "  const chips = SWEETENERS.map(t=>`\n"
        "    <button class='filter-btn sm green ${selectedSweeteners.has(t)?\"active\":\"\"}'"
        " onclick='toggleSweet(\"${t}\")'>${SWEETENER_LABELS[t]||t}</button>`).join('');\n"
        "  document.getElementById('sweetFilters').innerHTML = chips ?\n"
        "    (logicChip('sw', swLogic) + chips) :\n"
        "    `<span style='font-size:11px;color:#aaa'>aucun détecté</span>`;\n"
        "}\n"
        "function toggleSweet(t){toggleSet(selectedSweeteners,t);buildSweetFilters();updateTable();updateCharts();}\n"
        "function buildWheyFilters(){\n"
        "  const chips = WHEY_TYPES.map(t=>`\n"
        "    <button class='filter-btn sm purple ${selectedWheyTypes.has(t)?\"active\":\"\"}'"
        " onclick='toggleWhey(\"${t}\")'>${WHEY_TYPE_LABELS[t]||t}</button>`).join('');\n"
        "  document.getElementById('wheyFilters').innerHTML = chips ?\n"
        "    (logicChip('wt', wtLogic) + chips) :\n"
        "    `<span style='font-size:11px;color:#aaa'>aucun détecté</span>`;\n"
        "}\n"
        "function toggleWhey(t){toggleSet(selectedWheyTypes,t);buildWheyFilters();updateTable();updateCharts();}\n"
        "function getAllLabels(){\n"
        "  const s=new Set();\n"
        "  for(const p of PRODUCTS){ for(const lb of (effectiveTags(p).labels||[])) s.add(lb); }\n"
        "  return [...s].sort();\n"
        "}\n"
        "function buildLabelFilters(){\n"
        "  const labels=getAllLabels();\n"
        "  const row=document.getElementById('labelFiltersRow');\n"
        "  if(!labels.length){ row.style.display='none'; selectedLabels.clear(); return; }\n"
        "  row.style.display='flex';\n"
        "  const chips = labels.map(t=>`\n"
        "    <button class='filter-btn sm amber ${selectedLabels.has(t)?\"active\":\"\"}'"
        " onclick='toggleLabel(${escapeHtml(JSON.stringify(t))})'>${escapeHtml(t)}</button>`).join('');\n"
        "  document.getElementById('labelFilters').innerHTML = labels.length>1 ? (logicChip('lbl', lblLogic) + chips) : chips;\n"
        "}\n"
        "function toggleLabel(t){toggleSet(selectedLabels,t);buildLabelFilters();updateTable();updateCharts();}\n"
        # ── Modal d'édition ─────────────────────────────────────────────────
        "function openEditModal(produit){\n"
        "  modalCurrent = produit;\n"
        "  const r = RAW.find(x=>x.produit===produit);\n"
        "  const eff = effectiveTags(produit);\n"
        "  document.getElementById('modalProduit').textContent = produit;\n"
        "  const sw = (r&&r.sweeteners||[]).map(t=>SWEETENER_LABELS[t]||t).join(', ') || '—';\n"
        "  const wt = (r&&r.wheyTypes||[]).map(t=>WHEY_TYPE_LABELS[t]||t).join(', ') || '—';\n"
        "  const ingr = r ? (r.ingredients||'').slice(0,180) : '';\n"
        "  document.getElementById('modalAuto').innerHTML =\n"
        "    `<b>Détecté auto :</b><br>· Édulcorants : ${sw}<br>· Type protéine : ${wt}<br>` +\n"
        "    (ingr ? `<br><span style='color:#999'>${escapeHtml(ingr)}${ingr.length>=180?'…':''}</span>` :\n"
        "     `<br><span style='color:#B23B3B'>⚠️ Aucun ingrédient capturé</span>`);\n"
        "  document.getElementById('modalLabels').value = (eff.labels||[]).join(', ');\n"
        "  document.getElementById('modalNote').value = eff.note || '';\n"
        "  document.getElementById('editModal').classList.add('open');\n"
        "}\n"
        "function closeEditModal(){ document.getElementById('editModal').classList.remove('open'); modalCurrent=null; }\n"
        "function saveEdit(){\n"
        "  if(!modalCurrent) return;\n"
        "  const labels = document.getElementById('modalLabels').value\n"
        "    .split(',').map(s=>s.trim()).filter(Boolean);\n"
        "  const note = document.getElementById('modalNote').value.trim();\n"
        "  const srv = SERVER_TAGS[modalCurrent];\n"
        "  const sameAsServer = srv && JSON.stringify(srv.labels||[])===JSON.stringify(labels) && (srv.note||'')===note;\n"
        "  const empty = !labels.length && !note;\n"
        "  if(empty && !srv){ delete LOCAL_TAGS[modalCurrent]; }\n"
        "  else if(sameAsServer){ delete LOCAL_TAGS[modalCurrent]; }\n"
        "  else { LOCAL_TAGS[modalCurrent] = {labels, note}; }\n"
        "  saveLocalTags(LOCAL_TAGS);\n"
        "  closeEditModal();\n"
        "  refreshAll();\n"
        "}\n"
        "function resetProductEdit(){\n"
        "  if(!modalCurrent) return;\n"
        "  if(!confirm('Réinitialiser tags & note de ce produit (revient à tags.json) ?')) return;\n"
        "  delete LOCAL_TAGS[modalCurrent];\n"
        "  saveLocalTags(LOCAL_TAGS);\n"
        "  closeEditModal();\n"
        "  refreshAll();\n"
        "}\n"
        "function refreshAll(){ updateEditStatus(); buildLabelFilters(); updateTable(); }\n"
        "function updateEditStatus(){\n"
        "  const n = Object.keys(LOCAL_TAGS).length;\n"
        "  document.getElementById('editStatus').textContent = `${n} modif${n>1?'s':''} locale${n>1?'s':''}`;\n"
        "  document.getElementById('exportBtn').disabled = false;\n"
        "}\n"
        "function buildExportPayload(){\n"
        "  // Fusion : SERVER_TAGS écrasés par LOCAL_TAGS\n"
        "  const merged = {};\n"
        "  for(const [k,v] of Object.entries(SERVER_TAGS)){\n"
        "    if((v.labels&&v.labels.length) || (v.note||'').length) merged[k] = {labels:[...(v.labels||[])], note:v.note||''};\n"
        "  }\n"
        "  for(const [k,v] of Object.entries(LOCAL_TAGS)){\n"
        "    if((v.labels&&v.labels.length) || (v.note||'').length) merged[k] = {labels:[...(v.labels||[])], note:v.note||''};\n"
        "    else delete merged[k];\n"
        "  }\n"
        "  return merged;\n"
        "}\n"
        "function exportTags(){\n"
        "  const payload = buildExportPayload();\n"
        "  const blob = new Blob([JSON.stringify(payload, null, 2)], {type:'application/json'});\n"
        "  const url = URL.createObjectURL(blob);\n"
        "  const a = document.createElement('a');\n"
        "  a.href = url; a.download = 'tags.json'; a.click();\n"
        "  setTimeout(()=>URL.revokeObjectURL(url), 1000);\n"
        "}\n"
        "function clearLocalEdits(){\n"
        "  const n = Object.keys(LOCAL_TAGS).length;\n"
        "  if(!n) return;\n"
        "  if(!confirm(`Supprimer ${n} modif(s) locale(s) ? Les valeurs reviendront à tags.json.`)) return;\n"
        "  LOCAL_TAGS = {};\n"
        "  saveLocalTags(LOCAL_TAGS);\n"
        "  refreshAll();\n"
        "}\n"
        "// Sort key initial : adapté au tab par défaut\n"
        "sortKey = (TAB_PRIMARY[currentTab]||TAB_PRIMARY.whey).sort;\n"
        "buildTypeTabs();applyTabUI();\n"
        "buildMetrics();buildCategoryFilters();buildFilters();\n"
        "buildSweetFilters();buildWheyFilters();buildTypeTagFilters();buildLabelFilters();\n"
        "initCharts();\n"
        "buildTrendSelect();renderTrendChips();buildTrendChart();\n"
        "renderTableHead();setupSearch();updateEditStatus();updateTable();\n"
        "</script>\n"
        "</body>\n"
        "</html>\n"
    )

    dash_path = EXCEL_PATH.parent / "whey_dashboard.html"
    dash_path.write_text(html, encoding="utf-8")
    print(f"Dashboard : {dash_path}")


# ── Main ──────────────────────────────────────────────────────────────────────
async def main():
    print(f"\n{'='*62}")
    print(f"  HSN Whey Tracker — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*62}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-setuid-sandbox"],
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        )
        # Bloque les ressources inutiles pour aller plus vite
        await context.route(
            "**/*",
            lambda route: route.abort()
            if route.request.resource_type in {"image", "font", "media"}
            else route.continue_(),
        )
        page = await context.new_page()

        print("\nCollecte des URLs produits...")
        product_urls = await get_product_urls(page, CATEGORY_URLS)
        # Inclure les URLs supplémentaires + oméga-3 + créatine (déduplication)
        seen = {u.split("?")[0].rstrip("/") for u in product_urls}
        for u in EXTRA_URLS + OMEGA3_URLS + CREATINE_URLS:
            if u.split("?")[0].rstrip("/") not in seen:
                product_urls.append(u)
                seen.add(u.split("?")[0].rstrip("/"))
        print(f"\n  {len(product_urls)} produits uniques\n")
        await page.close()

        sem = asyncio.Semaphore(CONCURRENCY)
        total = len(product_urls)

        async def worker(idx: int, url: str):
            async with sem:
                short = url.split("/")[-1]
                print(f"[{idx:02d}/{total}] {short}")
                last_err = None
                for attempt in range(RETRY_ATTEMPTS + 1):
                    worker_page = await context.new_page()
                    try:
                        rows = await scrape_product(worker_page, url)
                        if rows:
                            return rows
                        last_err = "résultat vide"
                    except Exception as e:
                        last_err = repr(e)
                    finally:
                        await worker_page.close()
                    if attempt < RETRY_ATTEMPTS:
                        print(f"   ↻ retry {short} ({last_err})")
                        await asyncio.sleep(RETRY_DELAY_MS / 1000)
                log_error(url, f"Échec après {RETRY_ATTEMPTS + 1} tentative(s) : {last_err}")
                return []

        batches = await asyncio.gather(
            *(worker(i, u) for i, u in enumerate(product_urls, 1))
        )
        all_rows = [r for batch in batches for r in batch]

        await browser.close()

    if all_rows:
        print(f"\nEnregistrement de {len(all_rows)} lignes dans Excel...")
        append_rows(all_rows)
        print(f"Excel OK : {EXCEL_PATH}")
        print("Generation du dashboard HTML...")
        generate_dashboard()
        print(f"Termine ! {len(all_rows)} entrees pour le {date.today().isoformat()}")
    else:
        print("Aucune donnee collectee.")

    return all_rows


if __name__ == "__main__":
    asyncio.run(main())
